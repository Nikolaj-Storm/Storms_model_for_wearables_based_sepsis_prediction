# Copyright (c) 2026 Nikolaj Storm Petersen. Licensed under CC BY-NC 4.0.
# Non-commercial use only. If you use or adapt this code, please cite the author.
# See LICENSE and CITATION.cff  |  https://creativecommons.org/licenses/by-nc/4.0/

# ============================================================================
#  run_toks21_proxy.py
#  Stage: 5 - Comparators
#
#  PURPOSE
#    Strict Region Nordjylland TOKS 2.1 proxy on the 4-hour cohort. Computes the
#    per-variable TOKS scores with the protocol oxygen modifier and folds a
#    single-parameter red-flag override into an effective score, then sweeps
#    cutoffs and evaluates against the 6-hour and 12-hour Sepsis-3 labels. New
#    rows are inserted into the 4-hour block of the Excel workbook.
#
#  INPUTS
#    /PATH/TO/PROJECT/technical/Data/v3_dataset_4h_test_with_GCS_O2.parquet
#  OUTPUTS
#    /PATH/TO/PROJECT/technical/Data/v3_dataset_4h_test_TOKS21.parquet
#    /PATH/TO/PROJECT/all results updated.xlsx  (rows appended in 4-hour block)
#    Also prints sweep tables to the console.
#
#  USER-EDITABLE SETTINGS  (grep the body for the tag  EDIT:  to find each)
#    INPUT_PARQUET     -  test parquet to score
#    OUTPUT_PARQUET    -  per-row TOKS 2.1 score parquet to write
#    EXCEL_PATH        -  results workbook the operating points are appended to
#    CLINICAL_CUTOFFS  -  escalation cutoffs [1, 6, 8, 10]
#    TOKS scoring thresholds  -  per-variable cutoffs in the score_* helpers
#                         (RR, SpO2, SBP, HR, GCS, Temp)
#    red-flag cutoffs  -  single-parameter extreme cutoffs in red_flag()
#    target columns    -  is_sepsis_6h and is_sepsis_12h label columns
#
#  REQUIRES: numpy, pandas, openpyxl
# ============================================================================
"""
TOKS 2.1 proxy benchmark — corrected to match Region Nordjylland TOKS 2.1
protocol (Godkendt 2025-02-26, Dokument-id REGNORD-1547845390-107).

Replaces the earlier NEWS2-thresholded proxy with the actual Danish TOKS 2.1
variable thresholds, the protocol-correct oxygen modifier (added to SpO2
score), and protocol-documented escalation cutoffs at ≥1, ≥6, ≥8, ≥10. Also
reports a single-parameter red-flag variant for sensitivity analysis.

Outputs:
  - /PATH/TO/PROJECT/technical/Data/v3_dataset_4h_test_TOKS21.parquet
  - Updated rows appended to /PATH/TO/PROJECT/all results updated.xlsx

Run:  python3 run_toks21_proxy.py
"""

from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
# EDIT: input/output/workbook paths
INPUT_PARQUET = Path("/PATH/TO/PROJECT/technical/Data/v3_dataset_4h_test_with_GCS_O2.parquet")
OUTPUT_PARQUET = Path("/PATH/TO/PROJECT/technical/Data/v3_dataset_4h_test_TOKS21.parquet")
EXCEL_PATH = Path("/PATH/TO/PROJECT/all results updated.xlsx")

# EDIT: TOKS 2.1 clinical escalation cutoffs. Full integer sweep 0..max also computed.
CLINICAL_CUTOFFS = [1, 6, 8, 10]
CUTOFF_LABELS = {
    1: "Threshold >= 1 (Grøn, heightened attention)",
    6: "Threshold >= 6 (Gul, on-call physician 1h)",
    8: "Threshold >= 8 (Orange, urgent physician 15min)",
    10: "Threshold >= 10 (Rød, acute / continuous)",
}


# ---------------------------------------------------------------------------
# TOKS 2.1 variable scoring
# ---------------------------------------------------------------------------

def score_resprate(x):
    """Respiratory rate (per minute) per TOKS 2.1."""
    # EDIT: RR scoring thresholds
    if pd.isna(x):
        return 0
    if x <= 7: return 3
    if x <= 11: return 1
    if x <= 20: return 0
    if x <= 24: return 2
    return 3  # 25+


def score_spo2(x, on_o2):
    """SpO2 standard scale per TOKS 2.1, plus +2 if on supplemental oxygen."""
    # EDIT: SpO2 scoring thresholds (+2 supplemental-oxygen modifier)
    if pd.isna(x):
        return 0
    if x >= 96: base = 0
    elif x >= 94: base = 1  # 94, 95
    elif x >= 92: base = 2  # 92, 93
    elif x >= 90: base = 3  # 90, 91
    elif x >= 80: base = 3  # 80-89
    else: base = 3  # ≤79 red
    if on_o2 == 1:
        base += 2
    return base


def score_sbp(x):
    """Systolic BP per TOKS 2.1."""
    # EDIT: SBP scoring thresholds
    if pd.isna(x):
        return 0
    if x >= 220: return 3
    if x >= 110: return 0  # 110-219
    if x >= 100: return 1  # 100-109
    if x >= 90: return 2   # 90-99
    if x >= 80: return 3   # 80-89
    return 3               # ≤79 red


def score_hr(x):
    """Heart rate per TOKS 2.1."""
    # EDIT: HR scoring thresholds
    if pd.isna(x):
        return 0
    if x >= 141: return 3   # ≥141 red
    if x >= 130: return 3   # 130-140
    if x >= 111: return 2   # 111-129 (and 121-129)
    if x >= 110: return 2   # 110
    if x >= 90: return 1    # 90-109
    if x >= 50: return 0    # 50-89
    if x >= 40: return 1    # 40-49
    return 3                # ≤39 red


def score_gcs(x):
    """GCS per TOKS 2.1."""
    # EDIT: GCS scoring thresholds
    if pd.isna(x):
        return 0
    if x == 15: return 0
    return 3  # 14, 9-13, ≤8 all score 3


def score_temp(x):
    """Body temperature (°C) per TOKS 2.1."""
    # EDIT: Temperature scoring thresholds
    if pd.isna(x):
        return 0
    if x >= 40.1: return 2
    if x >= 39: return 2     # 39-40
    if x >= 38.1: return 1   # 38.1-38.9
    if x >= 38: return 1     # 38
    if x >= 36: return 0     # 36-37.9
    if x >= 35: return 1     # 35-35.9
    if x >= 34: return 3     # 34-34.9
    if x >= 32: return 3     # 32-33.9
    return 3                 # ≤31.9 red


def red_flag(x_resp, x_spo2, x_sbp, x_hr, x_gcs, x_temp):
    """Single-parameter red colour code per TOKS 2.1 protocol.

    Identified by column-3 ("B-justering, alarm ved ekstrem værdi") = 3 in
    the protocol's threshold table. Per the protocol, any single value with
    red colour code triggers the action algorithm for TOKS score ≥10.
    """
    # EDIT: single-parameter red-flag extreme cutoffs
    if not pd.isna(x_resp) and (x_resp <= 7 or x_resp >= 36): return 1
    if not pd.isna(x_spo2) and x_spo2 <= 79: return 1
    if not pd.isna(x_sbp) and x_sbp <= 79: return 1
    if not pd.isna(x_hr) and (x_hr >= 141 or x_hr <= 39): return 1
    if not pd.isna(x_gcs) and x_gcs <= 8: return 1
    if not pd.isna(x_temp) and x_temp <= 31.9: return 1
    return 0


# ---------------------------------------------------------------------------
# Compute TOKS_total + flags
# ---------------------------------------------------------------------------

def build_toks(df: pd.DataFrame) -> pd.DataFrame:
    print(f"Loaded {len(df):,} rows. Computing TOKS 2.1 scores...")
    pts_resp = df["resprate"].apply(score_resprate)
    pts_spo2 = df.apply(lambda r: score_spo2(r["spo2"], r["on_O2"]), axis=1)
    pts_sbp = df["sbp"].apply(score_sbp)
    pts_hr = df["heart_rate"].apply(score_hr)
    pts_gcs = df["GCS_total"].apply(score_gcs)
    pts_temp = df["temp_c"].apply(score_temp)

    toks_total = pts_resp + pts_spo2 + pts_sbp + pts_hr + pts_gcs + pts_temp
    rf = df.apply(lambda r: red_flag(r["resprate"], r["spo2"],
                                     r["sbp"], r["heart_rate"], r["GCS_total"],
                                     r["temp_c"]), axis=1)

    # EDIT: red-flag override target tier (red-flag rows promoted to >= 10)
    effective = np.maximum(toks_total.astype(int), np.where(rf == 1, 10, 0))

    has_missing_vital = df[["resprate", "spo2", "sbp", "heart_rate",
                            "GCS_total", "temp_c"]].isna().any(axis=1).astype(int)

    out = pd.DataFrame({
        "TOKS21_total": toks_total.astype(int),
        "TOKS21_effective": effective.astype(int),
        "red_flag": rf.astype(int),
        "has_missing_vital": has_missing_vital,
        "is_sepsis_6h": df["is_sepsis_6h"].fillna(0).astype(int),
        "is_sepsis_12h": df["is_sepsis_12h"].fillna(0).astype(int),
        "pts_resprate": pts_resp.astype(int),
        "pts_spo2": pts_spo2.astype(int),
        "pts_sbp": pts_sbp.astype(int),
        "pts_hr": pts_hr.astype(int),
        "pts_gcs": pts_gcs.astype(int),
        "pts_temp": pts_temp.astype(int),
    })
    return out


# ---------------------------------------------------------------------------
# Evaluation
# ---------------------------------------------------------------------------

def metrics_at_cutoff(scores, labels, k):
    flagged = scores >= k
    tp = int(((flagged) & (labels == 1)).sum())
    fp = int(((flagged) & (labels == 0)).sum())
    fn = int(((~flagged) & (labels == 1)).sum())
    tn = int(((~flagged) & (labels == 0)).sum())
    pos = tp + fn
    neg = fp + tn
    total = pos + neg
    recall = tp / pos if pos else 0.0
    fpr = fp / neg if neg else 0.0
    fnp = fn / pos if pos else 0.0
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    f1_pos = (2 * tp) / (2 * tp + fp + fn) if (2 * tp + fp + fn) > 0 else 0.0
    f1_neg = (2 * tn) / (2 * tn + fn + fp) if (2 * tn + fn + fp) > 0 else 0.0
    f1_macro = (f1_pos + f1_neg) / 2
    if precision + recall > 0:
        f0_5 = (1.25 * precision * recall) / (0.25 * precision + recall)
        f2 = (5 * precision * recall) / (4 * precision + recall)
    else:
        f0_5 = 0.0
        f2 = 0.0
    return {
        "Precision (PPV)": precision,
        "F0.5": f0_5,
        "F1 Macro": f1_macro,
        "F2": f2,
        "FPR": fpr,
        "FNP": fnp,
        "TP": tp, "FN": fn, "FP": fp, "TN": tn,
        "Threshold": k,
        "Recall": recall,
        "F1 (pos class)": f1_pos,
        "TP %": 100 * tp / total if total else 0.0,
        "FN %": 100 * fn / total if total else 0.0,
        "FP %": 100 * fp / total if total else 0.0,
        "TN %": 100 * tn / total if total else 0.0,
    }


def auroc_auprc(scores, labels):
    """Empirical AUROC and AUPRC over the integer score range."""
    s = np.asarray(scores)
    y = np.asarray(labels)
    cutoffs = list(range(int(s.min()), int(s.max()) + 2))
    rows = [metrics_at_cutoff(s, y, k) for k in cutoffs]
    fpr = np.array([r["FPR"] for r in rows])
    tpr = np.array([r["Recall"] for r in rows])
    prec = np.array([r["Precision (PPV)"] for r in rows])

    # AUROC via trapezoid on (FPR ascending) curve. cutoffs are descending
    # so reverse to get FPR ascending.
    order = np.argsort(fpr)
    auroc = float(np.trapz(tpr[order], fpr[order]))

    # AUPRC via trapezoid on (Recall ascending) curve.
    order = np.argsort(tpr)
    auprc = float(np.trapz(prec[order], tpr[order]))
    return auroc, auprc


def evaluate(scores, labels, label_name):
    print(f"\n--- TOKS 2.1 sweep against {label_name} ---")
    auroc, auprc = auroc_auprc(scores, labels)
    print(f"AUROC: {auroc:.4f}, AUPRC: {auprc:.4f}")

    rows = []
    # Find F1-Macro maximising cutoff across the integer range.
    score_max = int(scores.max())
    f1m_best = (-1, -1.0)
    for k in range(0, score_max + 2):
        m = metrics_at_cutoff(scores, labels, k)
        if m["F1 Macro"] > f1m_best[1]:
            f1m_best = (k, m["F1 Macro"])
    f1m_cutoff = f1m_best[0]

    cutoffs_to_report = list(CLINICAL_CUTOFFS)
    if f1m_cutoff not in cutoffs_to_report:
        cutoffs_to_report.append(f1m_cutoff)

    for k in cutoffs_to_report:
        m = metrics_at_cutoff(scores, labels, k)
        m["AUPRC"] = auprc
        m["AUROC"] = auroc
        is_f1max = (k == f1m_cutoff)
        if is_f1max:
            m["_variant_label"] = f"Max F1-Macro (Threshold >= {k})"
        else:
            m["_variant_label"] = CUTOFF_LABELS.get(k, f"Threshold >= {k}")
        rows.append(m)
        print(f"  thr ≥{k:2d} ({m['_variant_label']}): TP={m['TP']:5d}, FP={m['FP']:6d}, "
              f"recall={m['Recall']:.4f}, FPR={m['FPR']:.4f}, F1m={m['F1 Macro']:.4f}")
    return rows, auroc, auprc


# ---------------------------------------------------------------------------
# Excel append
# ---------------------------------------------------------------------------

EXCEL_COLS = ["Version", "Model", "Variant / Operating Point", "AUPRC",
              "Precision (PPV)", "AUROC", "F0.5", "F1 Macro", "F2", "FPR",
              "FNP", "TP", "FN", "FP", "TN", "Threshold", "Recall",
              "F1 (pos class)", "TP %", "FN %", "FP %", "TN %",
              "Train Acc", "Test Acc"]


def find_sheet_block_end(ws, sheet_name):
    """Find the row index just after the last 4-HOUR RESOLUTION data row."""
    last_data_row = None
    in_4h_block = False
    for row in range(1, ws.max_row + 2):
        cell = ws.cell(row=row, column=1).value
        if cell and "4-HOUR" in str(cell).upper():
            in_4h_block = True
            continue
        if in_4h_block:
            if cell and ("HOUR" in str(cell).upper() or "MINUTE" in str(cell).upper()) and "4-HOUR" not in str(cell).upper():
                break
            if cell and str(cell).strip():
                last_data_row = row
    return last_data_row


def append_to_excel(rows_6h, rows_12h, model_label):
    print(f"\nAppending to {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    for sheet_name, rows in [("6h Sepsis Prediction", rows_6h),
                             ("12h Sepsis Prediction", rows_12h)]:
        ws = wb[sheet_name]
        end = find_sheet_block_end(ws, sheet_name)
        if end is None:
            print(f"  WARNING: could not locate 4-hour block end in {sheet_name}, appending at sheet end")
            end = ws.max_row
        write_row = end + 1

        for r in rows:
            data = {
                "Version": "Rule-Based Baseline",
                "Model": model_label,
                "Variant / Operating Point": r["_variant_label"],
                "AUPRC": r["AUPRC"],
                "Precision (PPV)": r["Precision (PPV)"],
                "AUROC": r["AUROC"],
                "F0.5": r["F0.5"],
                "F1 Macro": r["F1 Macro"],
                "F2": r["F2"],
                "FPR": r["FPR"],
                "FNP": r["FNP"],
                "TP": r["TP"], "FN": r["FN"], "FP": r["FP"], "TN": r["TN"],
                "Threshold": r["Threshold"],
                "Recall": r["Recall"],
                "F1 (pos class)": r["F1 (pos class)"],
                "TP %": r["TP %"], "FN %": r["FN %"], "FP %": r["FP %"], "TN %": r["TN %"],
                "Train Acc": "", "Test Acc": "",
            }
            for col_idx, col_name in enumerate(EXCEL_COLS, 1):
                ws.cell(row=write_row, column=col_idx, value=data.get(col_name, ""))
            write_row += 1
        print(f"  {sheet_name}: appended {len(rows)} rows starting at row {end + 1}")

    wb.save(EXCEL_PATH)
    print(f"Saved {EXCEL_PATH}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"Reading {INPUT_PARQUET}...")
    df = pd.read_parquet(INPUT_PARQUET)
    print(f"  cohort: {df.shape[0]:,} rows, {df.index.get_level_values('stay_id').nunique():,} stays")
    print(f"  is_sepsis_6h prevalence:  {df['is_sepsis_6h'].mean():.4f}")
    print(f"  is_sepsis_12h prevalence: {df['is_sepsis_12h'].mean():.4f}")

    toks = build_toks(df)
    toks.index = df.index

    print(f"\nTOKS21_total range: {toks['TOKS21_total'].min()}..{toks['TOKS21_total'].max()}")
    print(f"TOKS21_effective range: {toks['TOKS21_effective'].min()}..{toks['TOKS21_effective'].max()}")
    print(f"red_flag fraction:    {toks['red_flag'].mean():.4f}")

    print(f"\nSaving per-row scores to {OUTPUT_PARQUET}...")
    toks.to_parquet(OUTPUT_PARQUET)

    # Strict-protocol scoring uses the effective score, namely TOKS21_total
    # with the red-flag override folded in (red-flag rows promoted to ≥10).
    rows_6h, auroc6, auprc6 = evaluate(
        toks["TOKS21_effective"].to_numpy(),
        toks["is_sepsis_6h"].to_numpy(),
        "is_sepsis_6h"
    )
    rows_12h, auroc12, auprc12 = evaluate(
        toks["TOKS21_effective"].to_numpy(),
        toks["is_sepsis_12h"].to_numpy(),
        "is_sepsis_12h"
    )

    append_to_excel(rows_6h, rows_12h,
                    model_label="TOKS 2.1 Proxy (strict protocol, 4h cadence)")


if __name__ == "__main__":
    main()
