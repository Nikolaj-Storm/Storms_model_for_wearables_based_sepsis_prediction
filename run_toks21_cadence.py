# Copyright (c) 2026 Nikolaj Storm Petersen. Licensed under CC BY-NC 4.0.
# Non-commercial use only. If you use or adapt this code, please cite the author.
# See LICENSE and CITATION.cff  |  https://creativecommons.org/licenses/by-nc/4.0/

# ============================================================================
#  run_toks21_cadence.py
#  Stage: 5 - Comparators
#
#  PURPOSE
#    Apply a strict TOKS 2.1 protocol observation cadence to per-stay 4-hour
#    rows. Walks each stay in time order, only "scores" at the interval the
#    previous TOKS tier dictates, forward-fills the last score between checks,
#    then sweeps cutoffs and evaluates against the 6-hour and 12-hour Sepsis-3
#    labels. Appends operating points to the Excel workbook.
#
#  INPUTS
#    /PATH/TO/PROJECT/technical/Data/v3_dataset_4h_test_TOKS21.parquet
#      (per-row TOKS 2.1 scores produced by run_toks21_proxy.py)
#  OUTPUTS
#    /PATH/TO/PROJECT/technical/Data/v3_dataset_4h_test_TOKS21_cadence.parquet
#    /PATH/TO/PROJECT/all results updated.xlsx  (rows appended)
#    Also prints sweep tables to the console.
#
#  USER-EDITABLE SETTINGS  (grep the body for the tag  EDIT:  to find each)
#    INPUT_PARQUET     -  per-row TOKS 2.1 score parquet to read
#    OUTPUT_PARQUET    -  cadence-applied score parquet to write
#    EXCEL_PATH        -  results workbook the operating points are appended to
#    CLINICAL_CUTOFFS  -  escalation cutoffs [1, 6, 8, 10] (Gron/Gul/Orange/Rod)
#    cadence intervals -  required_interval_hours thresholds (score 0 -> 12h,
#                         1-5 -> 8h, 6-7 -> 4h, 8+ -> 4h grid floor)
#    target columns    -  is_sepsis_6h and is_sepsis_12h label columns
#
#  REQUIRES: numpy, pandas, openpyxl
# ============================================================================
"""
TOKS 2.1 with strict protocol cadence simulation.

Extends the strict-protocol TOKS 2.1 proxy by walking each stay's 4-hour rows
in chronological order and only scoring at intervals dictated by the previous
TOKS tier. Between observations the most recent score is forward-filled, which
mirrors what a clinician at the bedside has access to between scheduled
checks.

Cadence per protocol:
  - Last score 0:        next observation in 12h (skip 2 of 3 4h rows)
  - Last score 1-5 (Grøn): next observation in 8h (skip 1 of 2 4h rows)
  - Last score 6-7 (Gul):  next observation in 4h (every row)
  - Last score ≥8:        every row (protocol wants 15min/continuous, 4h grid is the floor)

Reads:  technical/Data/v3_dataset_4h_test_TOKS21.parquet
Writes: technical/Data/v3_dataset_4h_test_TOKS21_cadence.parquet
        appends new rows to all results updated.xlsx
"""

from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl

# EDIT: input/output/workbook paths
INPUT_PARQUET = Path("/PATH/TO/PROJECT/technical/Data/v3_dataset_4h_test_TOKS21.parquet")
OUTPUT_PARQUET = Path("/PATH/TO/PROJECT/technical/Data/v3_dataset_4h_test_TOKS21_cadence.parquet")
EXCEL_PATH = Path("/PATH/TO/PROJECT/all results updated.xlsx")

# EDIT: TOKS 2.1 clinical escalation cutoffs
CLINICAL_CUTOFFS = [1, 6, 8, 10]
CUTOFF_LABELS = {
    1: "Threshold >= 1 (Grøn, heightened attention)",
    6: "Threshold >= 6 (Gul, on-call physician 1h)",
    8: "Threshold >= 8 (Orange, urgent physician 15min)",
    10: "Threshold >= 10 (Rød, acute / continuous)",
}
EXCEL_COLS = ["Version", "Model", "Variant / Operating Point", "AUPRC",
              "Precision (PPV)", "AUROC", "F0.5", "F1 Macro", "F2", "FPR",
              "FNP", "TP", "FN", "FP", "TN", "Threshold", "Recall",
              "F1 (pos class)", "TP %", "FN %", "FP %", "TN %",
              "Train Acc", "Test Acc"]


def required_interval_hours(last_score: int) -> int:
    # EDIT: protocol cadence intervals keyed on the last observed TOKS tier
    if last_score == 0: return 12
    if last_score <= 5: return 8
    if last_score <= 7: return 4
    return 4  # 8+ wants sub-grid; 4h is the grid floor


def simulate_stay(scores: np.ndarray) -> np.ndarray:
    """Forward-fill scores per protocol cadence on a single stay's 4h rows.

    Each row index corresponds to a 4h step. Returns the forward-filled score
    each row would carry given protocol-cadence observations.
    """
    n = len(scores)
    out = np.empty(n, dtype=int)
    last_score = int(scores[0])
    last_obs_idx = 0
    out[0] = last_score
    for i in range(1, n):
        hours_since = (i - last_obs_idx) * 4
        required = required_interval_hours(last_score)
        if hours_since >= required:
            last_score = int(scores[i])
            last_obs_idx = i
        out[i] = last_score
    return out


def metrics_at_cutoff(scores, labels, k):
    flagged = scores >= k
    tp = int(((flagged) & (labels == 1)).sum())
    fp = int(((flagged) & (labels == 0)).sum())
    fn = int(((~flagged) & (labels == 1)).sum())
    tn = int(((~flagged) & (labels == 0)).sum())
    pos = tp + fn
    neg = fp + tn
    total = pos + neg
    recall = tp / pos if pos else 0.0
    fpr = fp / neg if neg else 0.0
    fnp = fn / pos if pos else 0.0
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    f1_pos = (2 * tp) / (2 * tp + fp + fn) if (2 * tp + fp + fn) > 0 else 0.0
    f1_neg = (2 * tn) / (2 * tn + fn + fp) if (2 * tn + fn + fp) > 0 else 0.0
    f1_macro = (f1_pos + f1_neg) / 2
    if precision + recall > 0:
        f0_5 = (1.25 * precision * recall) / (0.25 * precision + recall)
        f2 = (5 * precision * recall) / (4 * precision + recall)
    else:
        f0_5 = 0.0
        f2 = 0.0
    return {
        "Precision (PPV)": precision, "F0.5": f0_5, "F1 Macro": f1_macro,
        "F2": f2, "FPR": fpr, "FNP": fnp,
        "TP": tp, "FN": fn, "FP": fp, "TN": tn,
        "Threshold": k, "Recall": recall, "F1 (pos class)": f1_pos,
        "TP %": 100 * tp / total if total else 0.0,
        "FN %": 100 * fn / total if total else 0.0,
        "FP %": 100 * fp / total if total else 0.0,
        "TN %": 100 * tn / total if total else 0.0,
    }


def auroc_auprc(scores, labels):
    s = np.asarray(scores)
    y = np.asarray(labels)
    cutoffs = list(range(int(s.min()), int(s.max()) + 2))
    rows = [metrics_at_cutoff(s, y, k) for k in cutoffs]
    fpr = np.array([r["FPR"] for r in rows])
    tpr = np.array([r["Recall"] for r in rows])
    prec = np.array([r["Precision (PPV)"] for r in rows])
    order = np.argsort(fpr)
    auroc = float(np.trapz(tpr[order], fpr[order]))
    order = np.argsort(tpr)
    auprc = float(np.trapz(prec[order], tpr[order]))
    return auroc, auprc


def evaluate(scores, labels, label_name):
    print(f"\n--- TOKS 2.1 (strict cadence) sweep against {label_name} ---")
    auroc, auprc = auroc_auprc(scores, labels)
    print(f"AUROC: {auroc:.4f}, AUPRC: {auprc:.4f}")
    rows = []
    score_max = int(scores.max())
    f1m_best = (-1, -1.0)
    for k in range(0, score_max + 2):
        m = metrics_at_cutoff(scores, labels, k)
        if m["F1 Macro"] > f1m_best[1]:
            f1m_best = (k, m["F1 Macro"])
    f1m_cutoff = f1m_best[0]
    cutoffs_to_report = list(CLINICAL_CUTOFFS)
    if f1m_cutoff not in cutoffs_to_report:
        cutoffs_to_report.append(f1m_cutoff)
    for k in cutoffs_to_report:
        m = metrics_at_cutoff(scores, labels, k)
        m["AUPRC"] = auprc
        m["AUROC"] = auroc
        m["_variant_label"] = (f"Max F1-Macro (Threshold >= {k})"
                               if k == f1m_cutoff
                               else CUTOFF_LABELS.get(k, f"Threshold >= {k}"))
        rows.append(m)
        print(f"  thr ≥{k:2d}: TP={m['TP']:5d}, FP={m['FP']:6d}, "
              f"recall={m['Recall']:.4f}, FPR={m['FPR']:.4f}, F1m={m['F1 Macro']:.4f}")
    return rows, auroc, auprc


def append_to_excel(rows_6h, rows_12h, model_label):
    print(f"\nAppending to {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    for sheet_name, rows in [("6h Sepsis Prediction", rows_6h),
                             ("12h Sepsis Prediction", rows_12h)]:
        ws = wb[sheet_name]
        write_row = ws.max_row + 1
        for r in rows:
            data = {
                "Version": "Rule-Based Baseline",
                "Model": model_label,
                "Variant / Operating Point": r["_variant_label"],
                "AUPRC": r["AUPRC"], "Precision (PPV)": r["Precision (PPV)"],
                "AUROC": r["AUROC"], "F0.5": r["F0.5"], "F1 Macro": r["F1 Macro"],
                "F2": r["F2"], "FPR": r["FPR"], "FNP": r["FNP"],
                "TP": r["TP"], "FN": r["FN"], "FP": r["FP"], "TN": r["TN"],
                "Threshold": r["Threshold"], "Recall": r["Recall"],
                "F1 (pos class)": r["F1 (pos class)"],
                "TP %": r["TP %"], "FN %": r["FN %"], "FP %": r["FP %"], "TN %": r["TN %"],
                "Train Acc": "", "Test Acc": "",
            }
            for col_idx, col_name in enumerate(EXCEL_COLS, 1):
                ws.cell(row=write_row, column=col_idx, value=data.get(col_name, ""))
            write_row += 1
        print(f"  {sheet_name}: appended {len(rows)} rows")
    wb.save(EXCEL_PATH)


def main():
    print(f"Reading {INPUT_PARQUET}...")
    df = pd.read_parquet(INPUT_PARQUET)
    print(f"  {df.shape[0]:,} rows, {df.index.get_level_values('stay_id').nunique():,} stays")

    print("\nApplying protocol cadence per stay...")
    df = df.sort_index()
    cadence_scores = np.empty(len(df), dtype=int)
    obs_flag = np.zeros(len(df), dtype=int)
    pos = 0
    for stay_id, group in df.groupby(level="stay_id", sort=False):
        n = len(group)
        scores = group["TOKS21_effective"].to_numpy()
        out = simulate_stay(scores)
        cadence_scores[pos:pos + n] = out
        # Mark observation rows
        last_score = int(scores[0])
        last_obs_idx = 0
        obs_flag[pos] = 1
        for i in range(1, n):
            hours_since = (i - last_obs_idx) * 4
            required = required_interval_hours(last_score)
            if hours_since >= required:
                obs_flag[pos + i] = 1
                last_score = int(scores[i])
                last_obs_idx = i
        pos += n

    df["TOKS21_cadence_score"] = cadence_scores
    df["protocol_observed"] = obs_flag

    obs_rate = obs_flag.mean()
    print(f"  Observation rate, fraction of 4h rows where TOKS would actually be scored: {obs_rate:.4f}")
    print(f"  ({obs_flag.sum():,} of {len(df):,} rows actually observed under protocol cadence)")

    print(f"\nSaving cadence-applied scores to {OUTPUT_PARQUET}...")
    df[["TOKS21_total", "TOKS21_effective", "TOKS21_cadence_score",
        "protocol_observed", "red_flag", "is_sepsis_6h", "is_sepsis_12h"]].to_parquet(OUTPUT_PARQUET)

    rows_6h, _, _ = evaluate(
        df["TOKS21_cadence_score"].to_numpy(),
        df["is_sepsis_6h"].to_numpy(),
        "is_sepsis_6h",
    )
    rows_12h, _, _ = evaluate(
        df["TOKS21_cadence_score"].to_numpy(),
        df["is_sepsis_12h"].to_numpy(),
        "is_sepsis_12h",
    )
    append_to_excel(rows_6h, rows_12h,
                    model_label="TOKS 2.1 Proxy (strict protocol + cadence)")


if __name__ == "__main__":
    main()
