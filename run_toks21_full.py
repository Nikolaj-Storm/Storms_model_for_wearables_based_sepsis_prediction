# Copyright (c) 2026 Nikolaj Storm Petersen. Licensed under CC BY-NC 4.0.
# Non-commercial use only. If you use or adapt this code, please cite the author.
# See LICENSE and CITATION.cff  |  https://creativecommons.org/licenses/by-nc/4.0/

# ============================================================================
#  run_toks21_full.py
#  Stage: 5 - Comparators
#
#  PURPOSE
#    Full-feature strict TOKS 2.1 proxy. Computes the per-variable TOKS scores,
#    the supplemental-oxygen modifier, single-parameter red-flag overrides,
#    sudden-change MAT criteria, sustained-low-urine trigger, and a protocol
#    cadence simulation at the chosen resolution, then evaluates against the
#    6-hour and 12-hour Sepsis-3 labels and appends to the Excel workbook.
#
#  INPUTS
#    /PATH/TO/PROJECT/technical/Data/v3_dataset_<res>_test_full.parquet
#      where <res> is 1h or 15m (chosen by CLI arg)
#  OUTPUTS
#    /PATH/TO/PROJECT/technical/Data/v3_dataset_<res>_test_TOKS21_strict.parquet
#    /PATH/TO/PROJECT/all results updated.xlsx  (rows appended)
#    Also prints sweep tables to the console.
#
#  USER-EDITABLE SETTINGS  (grep the body for the tag  EDIT:  to find each)
#    CLI arg            -  resolution key, one of 1h | 15m
#    RESOLUTION_CONFIG  -  per-resolution input/output paths and rows_per_hour
#    EXCEL_PATH         -  results workbook the operating points are appended to
#    CLINICAL_CUTOFFS   -  escalation cutoffs [1, 6, 8, 10]
#    TOKS scoring thresholds  -  per-variable cutoffs in build_toks / the
#                          score_* helpers (RR, SpO2, SBP, HR, GCS, Temp)
#    red-flag cutoffs   -  single-parameter extreme cutoffs in red_flag_extreme
#                          and build_toks
#    sudden-change deltas  -  MAT criteria in compute_sudden_change (HR/SBP/
#                          SpO2/RR/GCS deltas)
#    cadence intervals  -  required_interval_rows thresholds (12h/8h/4h/15min)
#    target columns     -  is_sepsis_6h and is_sepsis_12h label columns
#
#  REQUIRES: numpy, pandas, openpyxl
# ============================================================================
"""
TOKS 2.1 strict-protocol proxy with full feature set:
  - TOKS 2.1 variable thresholds (Region Nordjylland 2025)
  - Supplemental-oxygen +2 modifier on SpO2
  - Single-parameter red-colour-code override
  - Sudden-change MAT criteria (HR, SBP, SpO2, RR, GCS deltas)
  - Sustained low urine output (4h <50 mL) as MAT trigger
  - Protocol cadence simulation (12h / 8h / 4h / 15min / continuous)
  - Evaluates against is_sepsis_6h AND is_sepsis_12h
  - Runs at any resolution (rows-per-hour configurable)

Usage:
    python3 run_toks21_full.py 1h
    python3 run_toks21_full.py 15m
"""

import sys
from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl

# EDIT: results workbook path
EXCEL_PATH = Path("/PATH/TO/PROJECT/all results updated.xlsx")
# EDIT: per-resolution input/output parquet paths and rows_per_hour
RESOLUTION_CONFIG = {
    "1h": {
        "input": Path("/PATH/TO/PROJECT/technical/Data/v3_dataset_1h_test_full.parquet"),
        "output": Path("/PATH/TO/PROJECT/technical/Data/v3_dataset_1h_test_TOKS21_strict.parquet"),
        "rows_per_hour": 1,
        "label": "1-hour",
    },
    "15m": {
        "input": Path("/PATH/TO/PROJECT/technical/Data/v3_dataset_15m_test_full.parquet"),
        "output": Path("/PATH/TO/PROJECT/technical/Data/v3_dataset_15m_test_TOKS21_strict.parquet"),
        "rows_per_hour": 4,
        "label": "15-minute",
    },
}

# EDIT: TOKS 2.1 clinical escalation cutoffs
CLINICAL_CUTOFFS = [1, 6, 8, 10]
CUTOFF_LABELS = {
    1: "Threshold >= 1 (Grøn)",
    6: "Threshold >= 6 (Gul)",
    8: "Threshold >= 8 (Orange)",
    10: "Threshold >= 10 (Rød)",
}


def score_resprate(x):
    # EDIT: RR scoring thresholds
    if pd.isna(x): return 0
    if x <= 7: return 3
    if x <= 11: return 1
    if x <= 20: return 0
    if x <= 24: return 2
    return 3


def score_spo2(x, on_o2):
    # EDIT: SpO2 scoring thresholds (+2 supplemental-oxygen modifier)
    if pd.isna(x): return 0
    if x >= 96: base = 0
    elif x >= 94: base = 1
    elif x >= 92: base = 2
    else: base = 3
    if on_o2 == 1:
        base += 2
    return base


def score_sbp(x):
    # EDIT: SBP scoring thresholds
    if pd.isna(x): return 0
    if x >= 220: return 3
    if x >= 110: return 0
    if x >= 100: return 1
    if x >= 90: return 2
    return 3


def score_hr(x):
    # EDIT: HR scoring thresholds
    if pd.isna(x): return 0
    if x >= 141: return 3
    if x >= 130: return 3
    if x >= 110: return 2
    if x >= 90: return 1
    if x >= 50: return 0
    if x >= 40: return 1
    return 3


def score_gcs(x):
    # EDIT: GCS scoring thresholds
    if pd.isna(x): return 0
    if x == 15: return 0
    return 3


def score_temp(x):
    # EDIT: Temperature scoring thresholds
    if pd.isna(x): return 0
    if x >= 40.1: return 2
    if x >= 39: return 2
    if x >= 38: return 1
    if x >= 36: return 0
    if x >= 35: return 1
    return 3


def red_flag_extreme(rr, spo2, sbp, hr, gcs, temp):
    # EDIT: single-parameter red-flag extreme cutoffs
    if not pd.isna(rr) and (rr <= 7 or rr >= 36): return 1
    if not pd.isna(spo2) and spo2 <= 79: return 1
    if not pd.isna(sbp) and sbp <= 79: return 1
    if not pd.isna(hr) and (hr >= 141 or hr <= 39): return 1
    if not pd.isna(gcs) and gcs <= 8: return 1
    if not pd.isna(temp) and temp <= 31.9: return 1
    return 0


# ---------------------------------------------------------------------------
# Sudden-change MAT criteria — computed per stay against previous row vitals
# ---------------------------------------------------------------------------
def compute_sudden_change(stay_df: pd.DataFrame) -> np.ndarray:
    """Return array of 0/1 sudden-change flags for each row in the stay."""
    n = len(stay_df)
    flags = np.zeros(n, dtype=int)
    if n < 2:
        return flags
    hr = stay_df["heart_rate"].to_numpy()
    sbp = stay_df["sbp"].to_numpy()
    spo2 = stay_df["spo2"].to_numpy()
    rr = stay_df["resprate"].to_numpy()
    gcs = stay_df["GCS_total"].to_numpy()
    on_o2 = stay_df["on_O2"].to_numpy()

    for i in range(1, n):
        prev_hr, cur_hr = hr[i - 1], hr[i]
        prev_sbp, cur_sbp = sbp[i - 1], sbp[i]
        prev_spo2, cur_spo2 = spo2[i - 1], spo2[i]
        prev_rr, cur_rr = rr[i - 1], rr[i]
        prev_gcs, cur_gcs = gcs[i - 1], gcs[i]
        cur_o2 = on_o2[i]

        # EDIT: HR sudden-change cutoff. prev in 40-110, current outside
        if (not pd.isna(prev_hr) and not pd.isna(cur_hr)
                and 40 <= prev_hr <= 110
                and (cur_hr < 40 or cur_hr > 110)):
            flags[i] = 1
            continue
        # EDIT: SBP sudden-change cutoff. prev ≥ 80, current < 80
        if (not pd.isna(prev_sbp) and not pd.isna(cur_sbp)
                and prev_sbp >= 80 and cur_sbp < 80):
            flags[i] = 1
            continue
        # EDIT: SpO2 sudden-change cutoff on room air. prev ≥ 90, current < 90
        if (not pd.isna(prev_spo2) and not pd.isna(cur_spo2)
                and prev_spo2 >= 90 and cur_spo2 < 90
                and cur_o2 == 0):
            flags[i] = 1
            continue
        # EDIT: RR sudden-change cutoff. |Δ| > 8
        if (not pd.isna(prev_rr) and not pd.isna(cur_rr)
                and abs(cur_rr - prev_rr) > 8):
            flags[i] = 1
            continue
        # EDIT: GCS sudden-drop cutoff. ≥ 2
        if (not pd.isna(prev_gcs) and not pd.isna(cur_gcs)
                and (prev_gcs - cur_gcs) >= 2):
            flags[i] = 1
            continue
    return flags


# ---------------------------------------------------------------------------
# Cadence simulation — walk per stay, "observe" only at protocol intervals
# ---------------------------------------------------------------------------
def required_interval_rows(last_score: int, rows_per_hour: int) -> int:
    # EDIT: protocol cadence intervals keyed on the last observed TOKS tier
    if last_score == 0: return int(round(12 * rows_per_hour))
    if last_score <= 5: return int(round(8 * rows_per_hour))
    if last_score <= 7: return int(round(4 * rows_per_hour))
    if last_score <= 9: return max(1, int(round(0.25 * rows_per_hour)))  # 15min
    return max(1, int(round(0.25 * rows_per_hour)))  # continuous, grid floor


def simulate_cadence(scores: np.ndarray, rows_per_hour: int) -> np.ndarray:
    n = len(scores)
    out = np.empty(n, dtype=int)
    last_score = int(scores[0])
    last_obs = 0
    out[0] = last_score
    for i in range(1, n):
        rows_since = i - last_obs
        required = required_interval_rows(last_score, rows_per_hour)
        if rows_since >= required:
            last_score = int(scores[i])
            last_obs = i
        out[i] = last_score
    return out


# ---------------------------------------------------------------------------
# Build TOKS scores
# ---------------------------------------------------------------------------
def build_toks(df: pd.DataFrame, rows_per_hour: int) -> pd.DataFrame:
    print(f"  Computing TOKS 2.1 scores for {len(df):,} rows (vectorised)...")
    rr = df["resprate"].to_numpy()
    spo2 = df["spo2"].to_numpy()
    sbp = df["sbp"].to_numpy()
    hr = df["heart_rate"].to_numpy()
    gcs = df["GCS_total"].to_numpy()
    temp = df["temp_c"].to_numpy()
    on_o2 = df["on_O2"].to_numpy()

    # EDIT: RR scoring thresholds. ≤7→3, 8-11→1, 12-20→0, 21-24→2, ≥25→3
    pts_resp = np.where(rr <= 7, 3,
                np.where(rr <= 11, 1,
                np.where(rr <= 20, 0,
                np.where(rr <= 24, 2, 3))))
    pts_resp = np.where(np.isnan(rr), 0, pts_resp).astype(int)

    # EDIT: SpO2 scoring thresholds. ≥96→0, 94-95→1, 92-93→2, ≤91→3, +2 if on O2
    pts_spo2_base = np.where(spo2 >= 96, 0,
                     np.where(spo2 >= 94, 1,
                     np.where(spo2 >= 92, 2, 3)))
    pts_spo2_base = np.where(np.isnan(spo2), 0, pts_spo2_base)
    pts_spo2 = (pts_spo2_base + np.where(on_o2 == 1, 2, 0)).astype(int)

    # EDIT: SBP scoring thresholds. ≥220→3, 110-219→0, 100-109→1, 90-99→2, ≤89→3
    pts_sbp = np.where(sbp >= 220, 3,
               np.where(sbp >= 110, 0,
               np.where(sbp >= 100, 1,
               np.where(sbp >= 90, 2, 3))))
    pts_sbp = np.where(np.isnan(sbp), 0, pts_sbp).astype(int)

    # EDIT: HR scoring thresholds. ≥130→3, 110-129→2, 90-109→1, 50-89→0, 40-49→1, ≤39→3
    pts_hr = np.where(hr >= 110, np.where(hr >= 130, 3, 2),
              np.where(hr >= 90, 1,
              np.where(hr >= 50, 0,
              np.where(hr >= 40, 1, 3))))
    pts_hr = np.where(np.isnan(hr), 0, pts_hr).astype(int)

    # EDIT: GCS scoring thresholds. 15→0, else→3
    pts_gcs = np.where(gcs == 15, 0, 3)
    pts_gcs = np.where(np.isnan(gcs), 0, pts_gcs).astype(int)

    # EDIT: Temperature scoring thresholds. ≥39→2, 38-38.9→1, 36-37.9→0, 35-35.9→1, ≤34.9→3
    pts_temp = np.where(temp >= 39, 2,
                np.where(temp >= 38, 1,
                np.where(temp >= 36, 0,
                np.where(temp >= 35, 1, 3))))
    pts_temp = np.where(np.isnan(temp), 0, pts_temp).astype(int)

    toks_total = (pts_resp + pts_spo2 + pts_sbp + pts_hr + pts_gcs + pts_temp).astype(int)

    # EDIT: single-parameter red-flag extreme cutoffs (any value in red zone)
    rf_extreme = (
        ((~np.isnan(rr)) & ((rr <= 7) | (rr >= 36))) |
        ((~np.isnan(spo2)) & (spo2 <= 79)) |
        ((~np.isnan(sbp)) & (sbp <= 79)) |
        ((~np.isnan(hr)) & ((hr >= 141) | (hr <= 39))) |
        ((~np.isnan(gcs)) & (gcs <= 8)) |
        ((~np.isnan(temp)) & (temp <= 31.9))
    ).astype(int)

    rf_urine = df["urine_low"].astype(int).to_numpy()

    print(f"  Computing sudden-change flags per stay...")
    df_sorted = df.sort_index()
    sudden = np.zeros(len(df_sorted), dtype=int)
    pos = 0
    for stay_id, stay_df in df_sorted.groupby(level="stay_id", sort=False):
        n = len(stay_df)
        sudden[pos:pos + n] = compute_sudden_change(stay_df)
        pos += n

    # EDIT: red-flag override target tier (red-flag rows promoted to >= 10)
    any_redflag = ((rf_extreme == 1) | (rf_urine == 1) | (sudden == 1))
    effective = np.where(any_redflag, np.maximum(toks_total, 10), toks_total)

    print(f"  Applying protocol cadence simulation (rows_per_hour={rows_per_hour})...")
    cadence_score = np.empty(len(df_sorted), dtype=int)
    pos = 0
    for stay_id, stay_df in df_sorted.groupby(level="stay_id", sort=False):
        n = len(stay_df)
        # effective scores for this stay in chronological order
        stay_indices = stay_df.index
        # Get the integer position for each row of this stay in df_sorted
        eff_stay = effective[pos:pos + n]
        cadence_score[pos:pos + n] = simulate_cadence(eff_stay, rows_per_hour)
        pos += n

    out = pd.DataFrame({
        "TOKS21_total": toks_total,
        "rf_extreme": rf_extreme,
        "rf_urine": rf_urine,
        "rf_sudden_change": sudden,
        "TOKS21_effective": effective,
        "TOKS21_cadence_score": cadence_score,
        "is_sepsis_6h": df_sorted["is_sepsis_6h"].fillna(0).astype(int).values,
        "is_sepsis_12h": df_sorted["is_sepsis_12h"].fillna(0).astype(int).values,
    }, index=df_sorted.index)
    return out


# ---------------------------------------------------------------------------
# Evaluation
# ---------------------------------------------------------------------------
def metrics_at_cutoff(scores, labels, k):
    flagged = scores >= k
    tp = int(((flagged) & (labels == 1)).sum())
    fp = int(((flagged) & (labels == 0)).sum())
    fn = int(((~flagged) & (labels == 1)).sum())
    tn = int(((~flagged) & (labels == 0)).sum())
    pos = tp + fn
    neg = fp + tn
    total = pos + neg
    recall = tp / pos if pos else 0.0
    fpr = fp / neg if neg else 0.0
    fnp = fn / pos if pos else 0.0
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    f1_pos = (2 * tp) / (2 * tp + fp + fn) if (2 * tp + fp + fn) > 0 else 0.0
    f1_neg = (2 * tn) / (2 * tn + fn + fp) if (2 * tn + fn + fp) > 0 else 0.0
    f1_macro = (f1_pos + f1_neg) / 2
    if precision + recall > 0:
        f0_5 = (1.25 * precision * recall) / (0.25 * precision + recall)
        f2 = (5 * precision * recall) / (4 * precision + recall)
    else:
        f0_5 = 0.0
        f2 = 0.0
    return {
        "Precision (PPV)": precision, "F0.5": f0_5, "F1 Macro": f1_macro,
        "F2": f2, "FPR": fpr, "FNP": fnp,
        "TP": tp, "FN": fn, "FP": fp, "TN": tn,
        "Threshold": k, "Recall": recall, "F1 (pos class)": f1_pos,
        "TP %": 100 * tp / total if total else 0.0,
        "FN %": 100 * fn / total if total else 0.0,
        "FP %": 100 * fp / total if total else 0.0,
        "TN %": 100 * tn / total if total else 0.0,
    }


def auroc_auprc(scores, labels):
    cutoffs = list(range(int(scores.min()), int(scores.max()) + 2))
    rows = [metrics_at_cutoff(scores, labels, k) for k in cutoffs]
    fpr = np.array([r["FPR"] for r in rows])
    tpr = np.array([r["Recall"] for r in rows])
    prec = np.array([r["Precision (PPV)"] for r in rows])
    order = np.argsort(fpr)
    auroc = float(np.trapz(tpr[order], fpr[order]))
    order = np.argsort(tpr)
    auprc = float(np.trapz(prec[order], tpr[order]))
    return auroc, auprc


def evaluate(scores, labels, label_name):
    print(f"\n  Sweep against {label_name}")
    auroc, auprc = auroc_auprc(scores, labels)
    print(f"    AUROC {auroc:.4f}, AUPRC {auprc:.4f}")
    rows = []
    score_max = int(scores.max())
    f1m_best = (-1, -1.0)
    for k in range(0, score_max + 2):
        m = metrics_at_cutoff(scores, labels, k)
        if m["F1 Macro"] > f1m_best[1]:
            f1m_best = (k, m["F1 Macro"])
    f1m_cutoff = f1m_best[0]
    cutoffs_to_report = list(CLINICAL_CUTOFFS)
    if f1m_cutoff not in cutoffs_to_report:
        cutoffs_to_report.append(f1m_cutoff)
    for k in cutoffs_to_report:
        m = metrics_at_cutoff(scores, labels, k)
        m["AUPRC"] = auprc
        m["AUROC"] = auroc
        m["_variant_label"] = (f"Max F1-Macro (Threshold >= {k})"
                               if k == f1m_cutoff
                               else CUTOFF_LABELS.get(k, f"Threshold >= {k}"))
        rows.append(m)
        print(f"    thr ≥{k:2d}: TP={m['TP']:5d}, FP={m['FP']:7d}, "
              f"recall={m['Recall']:.4f}, FPR={m['FPR']:.4f}, F1m={m['F1 Macro']:.4f}")
    return rows


# ---------------------------------------------------------------------------
# Excel append
# ---------------------------------------------------------------------------
EXCEL_COLS = ["Version", "Model", "Variant / Operating Point", "AUPRC",
              "Precision (PPV)", "AUROC", "F0.5", "F1 Macro", "F2", "FPR",
              "FNP", "TP", "FN", "FP", "TN", "Threshold", "Recall",
              "F1 (pos class)", "TP %", "FN %", "FP %", "TN %",
              "Train Acc", "Test Acc"]


def append_to_excel(rows, sheet_name, model_label):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[sheet_name]
    write_row = ws.max_row + 1
    for r in rows:
        data = {
            "Version": "Rule-Based Baseline",
            "Model": model_label,
            "Variant / Operating Point": r["_variant_label"],
            "AUPRC": r["AUPRC"], "Precision (PPV)": r["Precision (PPV)"],
            "AUROC": r["AUROC"], "F0.5": r["F0.5"], "F1 Macro": r["F1 Macro"],
            "F2": r["F2"], "FPR": r["FPR"], "FNP": r["FNP"],
            "TP": r["TP"], "FN": r["FN"], "FP": r["FP"], "TN": r["TN"],
            "Threshold": r["Threshold"], "Recall": r["Recall"],
            "F1 (pos class)": r["F1 (pos class)"],
            "TP %": r["TP %"], "FN %": r["FN %"], "FP %": r["FP %"], "TN %": r["TN %"],
            "Train Acc": "", "Test Acc": "",
        }
        for col_idx, col_name in enumerate(EXCEL_COLS, 1):
            ws.cell(row=write_row, column=col_idx, value=data.get(col_name, ""))
        write_row += 1
    wb.save(EXCEL_PATH)


def main():
    if len(sys.argv) < 2 or sys.argv[1] not in RESOLUTION_CONFIG:
        print("Usage: python3 run_toks21_full.py [1h|15m]")
        sys.exit(1)
    res_key = sys.argv[1]
    cfg = RESOLUTION_CONFIG[res_key]

    print(f"\n=== TOKS 2.1 strict-protocol proxy at {cfg['label']} resolution ===")
    print(f"Reading {cfg['input']}...")
    df = pd.read_parquet(cfg["input"])
    print(f"  {df.shape[0]:,} rows, {df.index.get_level_values('stay_id').nunique():,} stays")
    print(f"  is_sepsis_6h prevalence:  {df['is_sepsis_6h'].mean():.4f}")
    print(f"  is_sepsis_12h prevalence: {df['is_sepsis_12h'].mean():.4f}")

    toks = build_toks(df, cfg["rows_per_hour"])

    print(f"\n  Red-flag extreme rate: {toks['rf_extreme'].mean():.4f}")
    print(f"  Red-flag urine rate:   {toks['rf_urine'].mean():.4f}")
    print(f"  Red-flag sudden rate:  {toks['rf_sudden_change'].mean():.4f}")
    print(f"  TOKS21_total range: {toks['TOKS21_total'].min()}..{toks['TOKS21_total'].max()}")
    print(f"  Effective score range: {toks['TOKS21_effective'].min()}..{toks['TOKS21_effective'].max()}")
    print(f"  Cadence score range:   {toks['TOKS21_cadence_score'].min()}..{toks['TOKS21_cadence_score'].max()}")

    print(f"\nSaving TOKS scores to {cfg['output']}...")
    toks.to_parquet(cfg["output"])

    # Evaluate against both targets at the cadence-applied score
    rows_6h = evaluate(toks["TOKS21_cadence_score"].to_numpy(),
                       toks["is_sepsis_6h"].to_numpy(),
                       "is_sepsis_6h")
    rows_12h = evaluate(toks["TOKS21_cadence_score"].to_numpy(),
                        toks["is_sepsis_12h"].to_numpy(),
                        "is_sepsis_12h")

    model_label = f"TOKS 2.1 Proxy (strict + cadence, {cfg['label']})"
    print(f"\nAppending results to {EXCEL_PATH}...")
    append_to_excel(rows_6h, "6h Sepsis Prediction", model_label)
    append_to_excel(rows_12h, "12h Sepsis Prediction", model_label)
    print("Done.")


if __name__ == "__main__":
    main()
