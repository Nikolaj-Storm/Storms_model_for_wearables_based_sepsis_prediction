# Copyright (c) 2026 Nikolaj Storm Petersen. Licensed under CC BY-NC 4.0.
# Non-commercial use only. If you use or adapt this code, please cite the author.
# See LICENSE and CITATION.cff  |  https://creativecommons.org/licenses/by-nc/4.0/

# ============================================================================
#  compile_excel.py
#  Stage: 4 - Modeling
#
#  PURPOSE
#    Compiles model-performance CSVs into a single formatted Excel workbook
#    with four sheets (full engineered, basic, raw vitals, optimised). Computes
#    F0.5/F2 columns as live Excel formulas, sorts rows by AUROC, and injects
#    two hardcoded V13b NOSE meta-learner rows for the 1-hour resolution.
#
#  INPUTS
#    ../Results/standard_algos_performance_all_engineered.csv
#    ../Results/guessing_baselines.csv
#    ../Results/basic_features_performance.csv   (optional)
#    ../Results/raw_vitals_performance.csv       (optional)
#    ../Results/optimised_performance.csv        (optional)
#  OUTPUTS
#    ../Results/Master_Performance_Comparison.xlsx
#
#  USER-EDITABLE SETTINGS  (grep the body for the tag  EDIT:  to find each)
#    Input CSV paths    -  the five ../Results/*.csv source files; assumes you
#                          run from the original technical/Models/ directory
#    excel_path         -  output workbook ../Results/Master_Performance_Comparison.xlsx
#    V13b LogReg row    -  hardcoded tp/fp/fn/fpr and ROC AUC / AUPRC literals
#    V13b XGBoost row   -  hardcoded tp2/fp2/fn2/fpr2 and ROC AUC / AUPRC literals
#    COLS               -  column order / headers written to each sheet
#    Resolution keys    -  ['15_min', '1_hour', '4_hour'] loop order
#
#  REQUIRES: openpyxl, pandas, numpy
# ============================================================================

import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Load source CSVs ────────────────────────────────────────────────────────
df_standard = pd.read_csv('../Results/standard_algos_performance_all_engineered.csv')  # EDIT: standard-algo results CSV
df_guess    = pd.read_csv('../Results/guessing_baselines.csv')  # EDIT: dummy/guessing baseline CSV

# Rename baseline rows
df_guess['Resolution'] = df_guess['Resolution'].apply(lambda x: x + ' (Dummy Classifier)')

# Manually add V13b results for 1_hour (LogReg Meta)
# EDIT: hardcoded V13b NOSE LogReg confusion-matrix counts and false-positive rate
tp  = 8926;  fp  = 226859; fn  = 3852;  fpr  = 0.3120
tn  = int((fp / fpr) - fp)
total  = tp + fp + fn + tn
recall  = tp / (tp + fn);  precision  = tp / (tp + fp)
f1  = 2 * (precision * recall) / (precision + recall)
v13b_lr = {
    'Resolution': '1_hour (V13b NOSE LogReg Meta)',
    'Train Accuracy (%)': 'N/A',
    'Test Accuracy (%)': f'{(tp+tn)/total*100:.2f}%',
    'Precision': f'{precision:.4f}', 'Recall': f'{recall:.4f}',
    'F1 Score': f'{f1:.4f}', 'ROC AUC': '0.7580', 'AUPRC': '0.0578',  # EDIT: hardcoded V13b LogReg AUROC / AUPRC
    'TP (Count)': tp, 'FN (Count)': fn, 'FP (Count)': fp, 'TN (Count)': tn,
    'TP (%)': f'{tp/total*100:.2f}%', 'FN (%)': f'{fn/total*100:.2f}%',
    'FP (%)': f'{fp/total*100:.2f}%', 'TN (%)': f'{tn/total*100:.2f}%',
}

# V13b XGBoost Meta
# EDIT: hardcoded V13b NOSE XGBoost confusion-matrix counts and false-positive rate
tp2 = 2234; fp2 = 23577; fn2 = 10544; fpr2 = 0.0324
tn2 = int((fp2 / fpr2) - fp2)
total2 = tp2 + fp2 + fn2 + tn2
recall2 = tp2 / (tp2 + fn2); precision2 = tp2 / (tp2 + fp2)
f1_2 = 2 * (precision2 * recall2) / (precision2 + recall2)
v13b_xgb = {
    'Resolution': '1_hour (V13b NOSE XGBoost Meta)',
    'Train Accuracy (%)': 'N/A',
    'Test Accuracy (%)': f'{(tp2+tn2)/total2*100:.2f}%',
    'Precision': f'{precision2:.4f}', 'Recall': f'{recall2:.4f}',
    'F1 Score': f'{f1_2:.4f}', 'ROC AUC': '0.7606', 'AUPRC': '0.0609',  # EDIT: hardcoded V13b XGBoost AUROC / AUPRC
    'TP (Count)': tp2, 'FN (Count)': fn2, 'FP (Count)': fp2, 'TN (Count)': tn2,
    'TP (%)': f'{tp2/total2*100:.2f}%', 'FN (%)': f'{fn2/total2*100:.2f}%',
    'FP (%)': f'{fp2/total2*100:.2f}%', 'TN (%)': f'{tn2/total2*100:.2f}%',
}

df_v13b = pd.DataFrame([v13b_lr, v13b_xgb])

# Load basic-features results (with time_since)
basic_csv = '../Results/basic_features_performance.csv'  # EDIT: basic-features results CSV (optional)
if os.path.exists(basic_csv):
    df_basic = pd.read_csv(basic_csv)
else:
    df_basic = pd.DataFrame()

# Load raw-vitals results (without time_since)
raw_csv = '../Results/raw_vitals_performance.csv'  # EDIT: raw-vitals results CSV (optional)
if os.path.exists(raw_csv):
    df_raw = pd.read_csv(raw_csv)
else:
    df_raw = pd.DataFrame()

# Load optimised results
opt_csv = '../Results/optimised_performance.csv'  # EDIT: optimised-features results CSV (optional)
if os.path.exists(opt_csv):
    df_opt = pd.read_csv(opt_csv)
else:
    df_opt = pd.DataFrame()

# ─── Helpers ─────────────────────────────────────────────────────────────────
COLS = ['Resolution', 'Train Accuracy (%)', 'Test Accuracy (%)',
        'Precision', 'Recall', 'F0.5 Score', 'F1 Score', 'F2 Score',
        'ROC AUC', 'AUPRC',
        'TP (Count)', 'FN (Count)', 'FP (Count)', 'TN (Count)',
        'TP (%)', 'FN (%)', 'FP (%)', 'TN (%)']  # EDIT: sheet column order / headers

COL_IDX = {c: i+1 for i, c in enumerate(COLS)}  # 1-based

def rows_for_resolution(res_key, df_src, df_base=None):
    """Return rows (as dicts) for a given resolution key, sorted best→worst AUROC."""
    std  = df_src[df_src['Resolution'].str.startswith(res_key)].copy()
    rows = [std]
    if df_base is not None:
        base = df_base[df_base['Resolution'].str.startswith(res_key)].copy()
        rows.append(base)
    combined = pd.concat(rows, ignore_index=True)
    combined['_is_dummy'] = combined['Resolution'].str.contains('Dummy Classifier')
    combined['_roc'] = pd.to_numeric(combined['ROC AUC'], errors='coerce').fillna(0)
    combined = combined.sort_values(['_is_dummy', '_roc'], ascending=[True, False])
    combined.drop(columns=['_is_dummy', '_roc'], inplace=True)
    return combined.to_dict('records')

def rows_for_resolution_full(res_key):
    """Full-feature rows: standard + V13b + dummy baseline."""
    std  = df_standard[df_standard['Resolution'].str.startswith(res_key)].copy()
    base = df_guess[df_guess['Resolution'].str.startswith(res_key)].copy()
    v13b = df_v13b[df_v13b['Resolution'].str.startswith(res_key)].copy() if res_key == '1_hour' else pd.DataFrame()
    combined = pd.concat([v13b, std, base], ignore_index=True)
    combined['_is_dummy'] = combined['Resolution'].str.contains('Dummy Classifier')
    combined['_roc'] = pd.to_numeric(combined['ROC AUC'], errors='coerce').fillna(0)
    combined = combined.sort_values(['_is_dummy', '_roc'], ascending=[True, False])
    combined.drop(columns=['_is_dummy', '_roc'], inplace=True)
    return combined.to_dict('records')

# ─── Style helpers ────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill('solid', fgColor='1F3864')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
DATA_FONT   = Font(size=10)
DUMMY_FILL  = PatternFill('solid', fgColor='F2F2F2')
ALT_FILL    = PatternFill('solid', fgColor='EBF3FB')
thin   = Side(style='thin', color='C0C0C0')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_header(ws, start_row):
    for col_i, col_name in enumerate(COLS, start=1):
        cell = ws.cell(row=start_row, column=col_i, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border    = BORDER

def f_formula(prec_col, rec_col, row, beta):
    """Excel F-beta formula. beta=0.5 → F0.5; beta=2 → F2."""
    b2 = beta**2
    return f'=(1+{b2})*{prec_col}{row}*{rec_col}{row}/({b2}*{prec_col}{row}+{rec_col}{row})'

def write_data_rows(ws, rows_list, start_row):
    prec_col = get_column_letter(COL_IDX['Precision'])
    rec_col  = get_column_letter(COL_IDX['Recall'])
    cur_row  = start_row
    for i, row_dict in enumerate(rows_list):
        is_dummy = 'Dummy Classifier' in row_dict.get('Resolution', '')
        fill = DUMMY_FILL if is_dummy else (ALT_FILL if i % 2 == 1 else None)
        for col_i, col_name in enumerate(COLS, start=1):
            if col_name == 'F0.5 Score':
                val  = f_formula(prec_col, rec_col, cur_row, 0.5)
                cell = ws.cell(row=cur_row, column=col_i, value=val)
            elif col_name == 'F2 Score':
                val  = f_formula(prec_col, rec_col, cur_row, 2)
                cell = ws.cell(row=cur_row, column=col_i, value=val)
            else:
                cell = ws.cell(row=cur_row, column=col_i, value=row_dict.get(col_name, ''))
            cell.font      = Font(bold=is_dummy, size=10, italic=is_dummy)
            if fill:
                cell.fill  = fill
            cell.alignment = Alignment(horizontal='center')
            cell.border    = BORDER
        cur_row += 1
    return cur_row  # next available row

def set_col_widths(ws):
    ws.column_dimensions['A'].width = 38
    for col_i in range(2, len(COLS)+1):
        ws.column_dimensions[get_column_letter(col_i)].width = 16

# ─── Build workbook ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Sheet 1: Full engineered features ────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Full Engineered Features'

current_row = 1
for res_key in ['15_min', '1_hour', '4_hour']:  # EDIT: resolution loop order
    write_header(ws1, current_row)
    data_rows = rows_for_resolution_full(res_key)
    current_row += 1
    current_row = write_data_rows(ws1, data_rows, current_row)
    current_row += 2  # blank separator rows

set_col_widths(ws1)
ws1.freeze_panes = 'B2'

# ── Sheet 2: Basic features ────────────────────────────────────────────────
ws2 = wb.create_sheet(title='Basic Features Only')

if df_basic.empty:
    ws2['A1'] = ('No results yet – run benchmark_basic_features.py first, '
                 'then re-run compile_excel.py.')
    ws2['A1'].font = Font(italic=True, color='FF0000')
else:
    current_row2 = 1
    for res_key in ['15_min', '1_hour', '4_hour']:
        write_header(ws2, current_row2)
        data_rows2 = rows_for_resolution(res_key, df_basic)
        current_row2 += 1
        current_row2 = write_data_rows(ws2, data_rows2, current_row2)
        current_row2 += 2

    set_col_widths(ws2)
    ws2.freeze_panes = 'B2'

# ── Sheet 3: Raw vitals only (no time_since) ──────────────────────────────
ws3 = wb.create_sheet(title='Raw Vitals Only')

if df_raw.empty:
    ws3['A1'] = ('No results yet – run benchmark_basic_features.py (with time_since removed) first, '
                 'then re-run compile_excel.py.')
    ws3['A1'].font = Font(italic=True, color='FF0000')
else:
    current_row3 = 1
    for res_key in ['15_min', '1_hour', '4_hour']:
        write_header(ws3, current_row3)
        data_rows3 = rows_for_resolution(res_key, df_raw)
        current_row3 += 1
        current_row3 = write_data_rows(ws3, data_rows3, current_row3)
        current_row3 += 2

    set_col_widths(ws3)
    ws3.freeze_panes = 'B2'

# ── Sheet 4: Optimised Features ───────────────────────────────────────────
ws4 = wb.create_sheet(title='Optimised Features')

if df_opt.empty:
    ws4['A1'] = 'No results yet – run optimise_features.py first, then re-run compile_excel.py.'
    ws4['A1'].font = Font(italic=True, color='FF0000')
else:
    current_row4 = 1
    for res_key in ['15_min', '1_hour', '4_hour']:
        write_header(ws4, current_row4)
        data_rows4 = rows_for_resolution(res_key, df_opt)
        current_row4 += 1
        current_row4 = write_data_rows(ws4, data_rows4, current_row4)
        current_row4 += 2

    set_col_widths(ws4)
    ws4.freeze_panes = 'B2'

# ─── Save ────────────────────────────────────────────────────────────────────
excel_path = '../Results/Master_Performance_Comparison.xlsx'  # EDIT: output workbook path
wb.save(excel_path)
print(f'Saved to {excel_path}')
