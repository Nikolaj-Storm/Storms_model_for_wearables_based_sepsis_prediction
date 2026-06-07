# Copyright (c) 2026 Nikolaj Storm Petersen. Licensed under CC BY-NC 4.0.
# Non-commercial use only. If you use or adapt this code, please cite the author.
# See LICENSE and CITATION.cff  |  https://creativecommons.org/licenses/by-nc/4.0/

# ============================================================================
#  generate_mimic_schema_excel.py
#  Stage: 1 - Data (schema documentation generator)
#
#  PURPOSE
#    Builds a multi-sheet Excel workbook documenting the MIMIC-IV v3.1
#    schema. For every table it writes a schema-definition sheet plus a
#    sheet of 10 deterministic mock sample rows, and adds dedicated sheets
#    for the V13b ML feature set and the raw vital signals used as inputs.
#
#  INPUTS
#    none (schema and sample values are hard-coded in this script)
#  OUTPUTS
#    /PATH/TO/PROJECT/technical/MIMIC_IV_Schema.xlsx
#
#  USER-EDITABLE SETTINGS  (grep the body for the tag  EDIT:  to find each)
#    random seed       -  random.seed(42) used twice for deterministic mock rows
#    output xlsx path  -  where the generated .xlsx workbook is written
#
#  REQUIRES: openpyxl
# ============================================================================
"""
MIMIC-IV Schema Excel Generator
Generates TWO sheets per table:
1. `[table_name] expl` - The schema definition (columns as rows)
2. `[table_name]` - The actual data snippet (columns as headers, 10 realistic sample rows)
Plain black & white theme logic.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import random

# ─── Style helpers ────────────────────────────────────────────────
def solid(hex_c):
    return PatternFill("solid", fgColor=hex_c)

THIN = Side(style="thin", color="000000")
B = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def f(bold=False, size=9, italic=False, color="000000"):
    return Font(bold=bold, name="Calibri", size=size, italic=italic, color=color)

def al(h="left"):
    return Alignment(horizontal=h, vertical="center", wrap_text=False)

HDR_FILL  = solid("1F1F1F")   # near-black header
COL_FILL  = solid("404040")   # dark grey
ALT_FILL  = solid("F2F2F2")   # light grey alternating rows
WHT_FILL  = solid("FFFFFF")   # white

def write_section_header(ws, row, text, ncols, size=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, name="Calibri", size=size, color="FFFFFF")
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = B
    ws.row_dimensions[row].height = 22

def write_col_headers(ws, row, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(bold=True, name="Calibri", size=9, color="FFFFFF")
        c.fill = COL_FILL
        c.alignment = al("center")
        c.border = B
    ws.row_dimensions[row].height = 16

def write_data_row(ws, row, values, col_aligns=None, alt=False):
    bg = ALT_FILL if alt else WHT_FILL
    for ci, v in enumerate(values, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = f(size=9)
        c.fill = bg
        align = (col_aligns[ci-1] if col_aligns else "left")
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(align == "left"))
        c.border = B
    ws.row_dimensions[row].height = 15

def auto_col_widths(ws, col_names, rows_data, min_w=8, max_w=40):
    for ci, name in enumerate(col_names, 1):
        max_len = len(str(name))
        for row in rows_data:
            if ci-1 < len(row):
                max_len = max(max_len, len(str(row[ci-1])))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_w, max(min_w, max_len + 2))

# ─── Mock Data Generator for 10 rows ──────────────────────────────
def generate_mock_rows(cols, n=10):
    rows = []
    random.seed(42) # Deterministic  # EDIT: random seed for mock sample rows

    # Base values
    base_vals = [c[5] for c in cols]

    for i in range(n):
        row = []
        for j, c in enumerate(cols):
            c_name = c[0]
            val = base_vals[j]
            dtype = c[1]

            if val == "null" or val is None:
                row.append("null" if i % 2 == 0 else "None")
                continue

            # Perturb IDs
            if c_name.endswith("_id") and isinstance(val, str) and val.isdigit():
                row.append(int(val) + i)
            elif dtype == "INT" or dtype == "SMALLINT":
                try:
                    num = int(val)
                    if c_name.endswith("_id") or "seq_num" in c_name:
                        row.append(num + i)
                    else:
                        row.append(num + random.randint(-1, 2))
                except:
                    row.append(val)
            elif dtype == "DOUBLE" or dtype == "REAL":
                try:
                    num = float(val)
                    row.append(round(num + random.uniform(-0.5, 0.5), 2))
                except:
                    row.append(val)
            elif dtype.startswith("TIMESTAMP") or dtype == "DATE":
                # For realistic looking data, just change the hour or minute string purely textually
                # Or keep it identical to save complexity
                try:
                    if len(val) > 10: # timestamp like 2180-05-06 22:23
                        # Increment minute by i*15
                        m = int(val[14:16])
                        new_m = (m + i*15) % 60
                        h = int(val[11:13]) + (m + i*15) // 60
                        row.append(f"{val[:11]}{h%24:02d}:{new_m:02d}")
                    else: # date
                        d = int(val[-2:])
                        new_d = d + i
                        row.append(f"{val[:-2]}{new_d:02d}")
                except:
                    row.append(val)
            else:
                # String properties
                if c_name == 'gender':
                    row.append(random.choice(['M', 'F']))
                else:
                    row.append(val)

        rows.append(row)
    return rows

# ─── Data Schema ──────────────────────────────────────────────────
schema = [
    # ══ HOSP ══
    {"module": "Hospital (hosp)", "table": "patients",
     "desc": "One row per unique patient.",
     "cols": [
         ("subject_id",        "INT",          "PK",  "",           "Unique patient identifier.",                         "10000032"),
         ("gender",            "VARCHAR(1)",   "",    "",           "Genotypical sex (M / F).",                           "M"),
         ("anchor_age",        "INT",          "",    "",           "Patient age at anchor year.",                        "52"),
         ("anchor_year",       "INT",          "",    "",           "Shifted calendar year of first admission.",          "2157"),
         ("anchor_year_group", "VARCHAR(255)", "",    "",           "3-year grouping for anchor year.",                   "2014-2016"),
         ("dod",               "TIMESTAMP",    "",    "",           "De-identified date of death (NULL if alive).",       "null"),
     ]},
    {"module": "Hospital (hosp)", "table": "admissions",
     "desc": "One row per hospital admission (hadm_id). A patient may have multiple admissions.",
     "cols": [
         ("subject_id",           "INT",          "FK", "patients",   "Patient identifier.",                              "10000032"),
         ("hadm_id",              "INT",          "PK", "",           "Unique hospital admission ID.",                    "22595853"),
         ("admittime",            "TIMESTAMP",    "",   "",           "Admission datetime.",                              "2180-05-06 22:23"),
         ("dischtime",            "TIMESTAMP",    "",   "",           "Discharge datetime.",                              "2180-05-07 14:00"),
         ("deathtime",            "TIMESTAMP",    "",   "",           "In-hospital death time (NULL if survived).",       "null"),
         ("admission_type",       "VARCHAR(40)",  "",   "",           "Urgency class: EW EMER., ELECTIVE, URGENT, …",    "EW EMER."),
         ("admission_location",   "VARCHAR(60)",  "",   "",           "Patient location before arrival.",                 "EMERGENCY ROOM"),
         ("discharge_location",   "VARCHAR(60)",  "",   "",           "Discharged to destination.",                      "HOME"),
         ("insurance",            "VARCHAR(255)", "",   "",           "Insurance type.",                                  "Medicare"),
         ("language",             "VARCHAR(10)",  "",   "",           "Primary language.",                                "ENGLISH"),
         ("marital_status",       "VARCHAR(30)",  "",   "",           "Marital status.",                                  "MARRIED"),
         ("race",                 "VARCHAR(80)",  "",   "",           "Self-reported race/ethnicity.",                    "WHITE"),
         ("edregtime",            "TIMESTAMP",    "",   "",           "ED registration datetime (NULL if not via ED).",  "2180-05-06 21:50"),
         ("edouttime",            "TIMESTAMP",    "",   "",           "ED departure datetime.",                           "2180-05-06 22:23"),
         ("hospital_expire_flag", "SMALLINT",     "",   "",           "1 = died in-hospital, 0 = survived.",             "0"),
     ]},
    {"module": "Hospital (hosp)", "table": "transfers",
     "desc": "Physical location/transfer events during a stay.",
     "cols": [
         ("subject_id",  "INT",          "FK", "patients",   "Patient identifier.",                            "10000032"),
         ("hadm_id",     "INT",          "FK", "admissions", "Admission identifier.",                          "22595853"),
         ("transfer_id", "INT",          "PK", "",           "Unique transfer event ID.",                      "8891"),
         ("eventtype",   "VARCHAR(10)",  "",   "",           "Event class: admit / transfer / discharge.",     "admit"),
         ("careunit",    "VARCHAR(255)", "",   "",           "Care unit name.",                                "MICU"),
         ("intime",      "TIMESTAMP",    "",   "",           "Arrival time at care unit.",                     "2180-05-06 22:23"),
         ("outtime",     "TIMESTAMP",    "",   "",           "Departure time from care unit.",                 "2180-05-07 14:00"),
     ]},
    {"module": "Hospital (hosp)", "table": "services",
     "desc": "Hospital service transitions (e.g. MED → MICU).",
     "cols": [
         ("subject_id",   "INT",         "FK", "patients",   "Patient identifier.",         "10000032"),
         ("hadm_id",      "INT",         "FK", "admissions", "Admission identifier.",       "22595853"),
         ("transfertime", "TIMESTAMP",   "",   "",           "Transfer time.",              "2180-05-06 22:23"),
         ("prev_service", "VARCHAR(20)", "",   "",           "Previous service (nullable).", "null"),
         ("curr_service", "VARCHAR(20)", "",   "",           "Current service.",            "MICU"),
     ]},
    {"module": "Hospital (hosp)", "table": "labevents",
     "desc": "Laboratory measurements (largest hosp table).",
     "cols": [
         ("labevent_id",    "INT",         "PK", "",          "Unique lab event ID.",                           "1734"),
         ("subject_id",     "INT",         "FK", "patients",  "Patient identifier.",                            "10000032"),
         ("hadm_id",        "INT",         "FK", "admissions","Admission identifier.",                          "22595853"),
         ("specimen_id",    "INT",         "",   "",          "Specimen collected.",                            "7777"),
         ("itemid",         "INT",         "FK", "d_labitems","Lab concept identifier.",                        "51006"),
         ("charttime",      "TIMESTAMP",   "",   "",          "Time measurement was taken.",                   "2180-05-06 22:45"),
         ("storetime",      "TIMESTAMP",   "",   "",          "Time result was stored in the system.",         "2180-05-06 23:10"),
         ("value",          "VARCHAR(200)","",   "",          "Raw string value.",                             "17.6"),
         ("valuenum",       "DOUBLE",      "",   "",          "Numeric value.",                                "17.6"),
         ("valueuom",       "VARCHAR(20)", "",   "",          "Unit of measure.",                              "mg/dL"),
         ("ref_range_lower","DOUBLE",      "",   "",          "Lower reference range.",                        "7.0"),
         ("ref_range_upper","DOUBLE",      "",   "",          "Upper reference range.",                        "25.0"),
         ("flag",           "VARCHAR(10)", "",   "",          "Abnormality flag (abnormal / delta).",          "abnormal"),
         ("priority",       "VARCHAR(7)",  "",   "",          "Order priority: ROUTINE or STAT.",              "STAT"),
         ("comments",       "TEXT",        "",   "",          "Free-text comments.",                           "null"),
     ]},
    {"module": "Hospital (hosp)", "table": "d_labitems",
     "desc": "Dictionary of lab item IDs.",
     "cols": [
         ("itemid",   "INT",        "PK", "", "Lab concept ID.",                                "51006"),
         ("label",    "VARCHAR(50)","",   "", "Human-readable label.",                          "Urea Nitrogen"),
         ("fluid",    "VARCHAR(50)","",   "", "Specimen fluid (Blood, Urine, …).",              "Blood"),
         ("category", "VARCHAR(50)","",   "", "Category (Chemistry, Hematology, ABG, …).",     "Chemistry"),
     ]},
    {"module": "Hospital (hosp)", "table": "diagnoses_icd",
     "desc": "ICD-9/10 diagnosis codes per admission.",
     "cols": [
         ("subject_id",  "INT",        "FK", "patients",        "Patient identifier.",             "10000032"),
         ("hadm_id",     "INT",        "FK", "admissions",      "Admission identifier.",           "22595853"),
         ("seq_num",     "INT",        "",   "",                "Priority rank (1 = primary).",    "1"),
         ("icd_code",    "VARCHAR(7)", "FK", "d_icd_diagnoses", "ICD diagnosis code.",             "A4151"),
         ("icd_version", "INT",        "",   "",                "ICD version (9 or 10).",          "10"),
     ]},
    {"module": "Hospital (hosp)", "table": "d_icd_diagnoses",
     "desc": "ICD code dictionary for diagnoses.",
     "cols": [
         ("icd_code",   "CHAR(7)",      "PK", "", "ICD code.",                    "A4151"),
         ("icd_version","INT",          "PK", "", "ICD version (9 or 10).",       "10"),
         ("long_title", "VARCHAR(255)", "",   "", "Full diagnosis description.",   "Septic shock"),
     ]},
    {"module": "Hospital (hosp)", "table": "procedures_icd",
     "desc": "ICD-coded procedures per admission.",
     "cols": [
         ("subject_id",  "INT",        "FK", "patients",          "Patient identifier.",       "10000032"),
         ("hadm_id",     "INT",        "FK", "admissions",        "Admission identifier.",     "22595853"),
         ("seq_num",     "INT",        "",   "",                  "Procedure priority order.", "1"),
         ("icd_code",    "VARCHAR(7)", "FK", "d_icd_procedures",  "ICD procedure code.",       "5A1935Z"),
         ("icd_version", "INT",        "",   "",                  "ICD version (9 or 10).",    "10"),
         ("chartdate",   "DATE",       "",   "",                  "Date performed.",           "2180-05-06"),
     ]},
    {"module": "Hospital (hosp)", "table": "d_icd_procedures",
     "desc": "ICD code dictionary for procedures.",
     "cols": [
         ("icd_code",   "CHAR(7)",      "PK", "", "ICD procedure code.",         "5A1935Z"),
         ("icd_version","INT",          "PK", "", "ICD version (9 or 10).",      "10"),
         ("long_title", "VARCHAR(255)", "",   "", "Full procedure description.", "Mechanical ventilation ≥96 hrs"),
     ]},
    {"module": "Hospital (hosp)", "table": "prescriptions",
     "desc": "Medication prescriptions per admission.",
     "cols": [
         ("subject_id",       "INT",          "FK", "patients",   "Patient identifier.",              "10000032"),
         ("hadm_id",          "INT",          "FK", "admissions", "Admission identifier.",            "22595853"),
         ("pharmacy_id",      "INT",          "FK", "pharmacy",   "Pharmacy record.",                 "4321"),
         ("starttime",        "TIMESTAMP",    "",   "",           "Prescription start.",              "2180-05-06 22:30"),
         ("stoptime",         "TIMESTAMP",    "",   "",           "Prescription end.",                "2180-05-07 06:00"),
         ("drug",             "VARCHAR(255)", "",   "",           "Drug name.",                       "Vancomycin"),
         ("drug_type",        "VARCHAR(20)",  "",   "",           "Drug type (BASE, ADDITIVE, …).",   "BASE"),
         ("dose_val_rx",      "VARCHAR(100)", "",   "",           "Prescribed dose value.",           "1000"),
         ("dose_unit_rx",     "VARCHAR(50)",  "",   "",           "Dose unit.",                       "mg"),
         ("doses_per_24_hrs", "REAL",         "",   "",           "Doses per 24 hours.",              "2"),
         ("route",            "VARCHAR(50)",  "",   "",           "Administration route.",            "IV"),
     ]},
    {"module": "Hospital (hosp)", "table": "pharmacy",
     "desc": "Pharmacy-level medication orders.",
     "cols": [
         ("subject_id",  "INT",       "FK", "patients",   "Patient identifier.",       "10000032"),
         ("hadm_id",     "INT",       "FK", "admissions", "Admission identifier.",     "22595853"),
         ("pharmacy_id", "INT",       "PK", "",           "Pharmacy order ID.",        "4321"),
         ("medication",  "TEXT",      "",   "",           "Medication name.",          "Vancomycin"),
         ("starttime",   "TIMESTAMP", "",   "",           "Order start time.",         "2180-05-06 22:30"),
         ("stoptime",    "TIMESTAMP", "",   "",           "Order end time.",           "2180-05-07 06:00"),
         ("status",      "VARCHAR",   "",   "",           "Order status.",             "Completed"),
         ("route",       "VARCHAR",   "",   "",           "Administration route.",     "IV"),
         ("frequency",   "VARCHAR",   "",   "",           "Dosing frequency.",         "Q12H"),
     ]},
    {"module": "Hospital (hosp)", "table": "emar",
     "desc": "Electronic Medication Administration Records.",
     "cols": [
         ("subject_id",  "INT",         "FK", "patients",   "Patient identifier.",         "10000032"),
         ("hadm_id",     "INT",         "FK", "admissions", "Admission identifier.",       "22595853"),
         ("emar_id",     "VARCHAR(25)", "PK", "",           "Unique EMAR record ID.",      "10000032-1"),
         ("emar_seq",    "INT",         "",   "",           "Sequence number.",            "1"),
         ("poe_id",      "VARCHAR(25)", "FK", "poe",        "Provider order entry ID.",    "10000032-1"),
         ("pharmacy_id", "INT",         "FK", "pharmacy",   "Pharmacy record.",            "4321"),
         ("charttime",   "TIMESTAMP",   "",   "",           "Administration time.",        "2180-05-06 23:00"),
         ("medication",  "TEXT",        "",   "",           "Drug administered.",          "Vancomycin"),
         ("event_txt",   "VARCHAR",     "",   "",           "Event description.",          "Administered"),
         ("storetime",   "TIMESTAMP",   "",   "",           "Time record was stored.",     "2180-05-06 23:05"),
     ]},
    {"module": "Hospital (hosp)", "table": "omr",
     "desc": "Outpatient measurement records (BP, weight, BMI, etc.).",
     "cols": [
         ("subject_id",   "INT",         "FK", "patients", "Patient identifier.",  "10000032"),
         ("chartdate",    "DATE",        "",   "",         "Measurement date.",    "2180-04-10"),
         ("seq_num",      "INT",         "",   "",         "Sequence number.",     "1"),
         ("result_name",  "VARCHAR(100)","",   "",         "Observation name.",    "Blood Pressure"),
         ("result_value", "TEXT",        "",   "",         "Observation value.",   "118/76"),
     ]},
    {"module": "Hospital (hosp)", "table": "poe",
     "desc": "Provider Order Entry — all clinical orders.",
     "cols": [
         ("poe_id",        "VARCHAR(25)", "PK", "",          "Unique order ID.",                     "10000032-1"),
         ("poe_seq",       "INT",         "",   "",          "Sequence within order.",               "1"),
         ("subject_id",    "INT",         "FK", "patients",  "Patient identifier.",                  "10000032"),
         ("hadm_id",       "INT",         "FK", "admissions","Admission identifier.",                "22595853"),
         ("ordertime",     "TIMESTAMP",   "",   "",          "Order timestamp.",                     "2180-05-06 22:30"),
         ("order_type",    "VARCHAR(25)", "",   "",          "Order type (Lab, Medications, …).",    "Medications"),
         ("order_subtype", "VARCHAR(50)", "",   "",          "Order sub-type.",                      "IV Antibiotic"),
         ("order_status",  "VARCHAR(15)", "",   "",          "Order status.",                        "Active"),
     ]},
    {"module": "Hospital (hosp)", "table": "drgcodes",
     "desc": "Diagnosis-related group billing codes per admission.",
     "cols": [
         ("subject_id",    "INT",          "FK", "patients",   "Patient identifier.",          "10000032"),
         ("hadm_id",       "INT",          "FK", "admissions", "Admission identifier.",        "22595853"),
         ("drg_type",      "VARCHAR(4)",   "",   "",           "DRG ontology (HCFA / APR).",   "APR"),
         ("drg_code",      "VARCHAR(10)",  "",   "",           "DRG code.",                    "137"),
         ("description",   "VARCHAR(195)", "",   "",           "DRG description.",             "Major resp. diagnoses"),
         ("drg_severity",  "SMALLINT",     "",   "",           "Severity level (1-4).",        "3"),
         ("drg_mortality", "SMALLINT",     "",   "",           "Mortality risk level (1-4).",  "2"),
     ]},
    {"module": "Hospital (hosp)", "table": "hcpcsevents",
     "desc": "HCPCS billing procedure codes per admission.",
     "cols": [
         ("subject_id",       "INT",         "FK", "patients",   "Patient identifier.",      "10000032"),
         ("hadm_id",          "INT",         "FK", "admissions", "Admission identifier.",    "22595853"),
         ("chartdate",        "DATE",        "",   "",           "Procedure date.",          "2180-05-06"),
         ("hcpcs_cd",         "CHAR(5)",     "FK", "d_hcpcs",    "HCPCS code.",              "G0378"),
         ("seq_num",          "INT",         "",   "",           "Sequence number.",         "1"),
         ("short_description","VARCHAR(180)","",   "",           "Procedure description.",   "Hospital obs per hr"),
     ]},
    {"module": "Hospital (hosp)", "table": "d_hcpcs",
     "desc": "Dictionary for HCPCS billing codes.",
     "cols": [
         ("code",             "CHAR(5)",     "PK", "", "Five-character HCPCS code.",   "G0378"),
         ("category",         "SMALLINT",    "",   "", "Broad category code.",         "4"),
         ("long_description", "TEXT",        "",   "", "Long textual description.",    "Hospital observation per hour"),
         ("short_description","VARCHAR(180)","",   "", "Short textual description.",   "Hospital obs per hr"),
     ]},
    {"module": "Hospital (hosp)", "table": "provider",
     "desc": "De-identified provider identifiers.",
     "cols": [
         ("provider_id", "VARCHAR(10)", "PK", "", "De-identified provider ID.", "PA6001"),
     ]},

    # ══ ICU ══
    {"module": "ICU (icu)", "table": "icustays",
     "desc": "One row per ICU stay. A patient may have multiple ICU stays per admission.",
     "cols": [
         ("subject_id",     "INT",       "FK", "patients",   "Patient identifier.",             "10000032"),
         ("hadm_id",        "INT",       "FK", "admissions", "Admission identifier.",           "22595853"),
         ("stay_id",        "INT",       "PK", "",           "Unique ICU stay ID.",             "30000026"),
         ("first_careunit", "VARCHAR",   "",   "",           "First ICU care unit.",            "MICU"),
         ("last_careunit",  "VARCHAR",   "",   "",           "Last ICU care unit.",             "MICU"),
         ("intime",         "TIMESTAMP", "",   "",           "ICU admission time.",             "2180-05-06 22:40"),
         ("outtime",        "TIMESTAMP", "",   "",           "ICU discharge time.",             "2180-05-13 09:15"),
         ("los",            "DOUBLE",    "",   "",           "ICU length of stay (days).",      "6.44"),
     ]},
    {"module": "ICU (icu)", "table": "chartevents",
     "desc": "High-frequency bedside charted observations (vitals, neuro assessments).",
     "cols": [
         ("stay_id",      "INT",       "FK", "icustays",  "ICU stay identifier.",           "30000026"),
         ("subject_id",   "INT",       "FK", "patients",  "Patient identifier.",            "10000032"),
         ("itemid",       "INT",       "FK", "d_items",   "Charted concept identifier.",    "220045"),
         ("charttime",    "TIMESTAMP", "",   "",          "Time chart was recorded.",       "2180-05-07 00:00"),
         ("storetime",    "TIMESTAMP", "",   "",          "Time result was stored.",        "2180-05-07 00:02"),
         ("value",        "VARCHAR",   "",   "",          "Raw string value.",              "88"),
         ("valuenum",     "DOUBLE",    "",   "",          "Numeric value.",                 "88.0"),
         ("valueuom",     "VARCHAR",   "",   "",          "Unit of measure.",               "bpm"),
         ("caregiver_id", "INT",       "FK", "caregiver", "De-identified caregiver ID.",    "22301"),
     ]},
    {"module": "ICU (icu)", "table": "d_items",
     "desc": "Dictionary of all item IDs used in ICU charting.",
     "cols": [
         ("itemid",       "INT",     "PK", "", "Item identifier.",                           "220045"),
         ("label",        "VARCHAR", "",   "", "Human-readable label.",                      "Heart Rate"),
         ("abbreviation", "VARCHAR", "",   "", "Short label.",                               "HR"),
         ("dhlevel",      "VARCHAR", "",   "", "Data hierarchy level.",                      "Routine Vital Signs"),
         ("category",     "VARCHAR", "",   "", "Clinical category.",                         "Routine Vital Signs"),
         ("unitname",     "VARCHAR", "",   "", "Default unit name.",                         "bpm"),
         ("param_type",   "VARCHAR", "",   "", "Parameter type (Numeric / Text / Date).",    "Numeric"),
     ]},
    {"module": "ICU (icu)", "table": "inputevents",
     "desc": "IV fluids and medication inputs administered in the ICU.",
     "cols": [
         ("stay_id",        "INT",       "FK", "icustays",  "ICU stay identifier.",         "30000026"),
         ("subject_id",     "INT",       "FK", "patients",  "Patient identifier.",          "10000032"),
         ("itemid",         "INT",       "FK", "d_items",   "Input item identifier.",       "220864"),
         ("charttime",      "TIMESTAMP", "",   "",          "Time of input.",               "2180-05-07 02:00"),
         ("amount",         "DOUBLE",    "",   "",          "Amount given.",                "500.0"),
         ("amountuom",      "VARCHAR",   "",   "",          "Amount unit.",                 "ml"),
         ("rate",           "DOUBLE",    "",   "",          "Rate of delivery.",            "125.0"),
         ("rateuom",        "VARCHAR",   "",   "",          "Rate unit.",                   "ml/hour"),
         ("totalamount",    "DOUBLE",    "",   "",          "Total cumulative amount.",     "2000.0"),
         ("totalamountuom", "VARCHAR",   "",   "",          "Total amount unit.",           "ml"),
         ("caregiver_id",   "INT",       "FK", "caregiver", "De-identified caregiver ID.",  "22301"),
     ]},
    {"module": "ICU (icu)", "table": "outputevents",
     "desc": "Fluid outputs (urine, drains, etc.) measured in the ICU.",
     "cols": [
         ("stay_id",      "INT",       "FK", "icustays",  "ICU stay identifier.",        "30000026"),
         ("subject_id",   "INT",       "FK", "patients",  "Patient identifier.",         "10000032"),
         ("itemid",       "INT",       "FK", "d_items",   "Output item identifier.",     "226559"),
         ("charttime",    "TIMESTAMP", "",   "",          "Time of output.",             "2180-05-07 06:00"),
         ("value",        "DOUBLE",    "",   "",          "Output volume.",              "250.0"),
         ("valueuom",     "VARCHAR",   "",   "",          "Volume unit.",                "ml"),
         ("caregiver_id", "INT",       "FK", "caregiver", "De-identified caregiver ID.", "22301"),
     ]},
    {"module": "ICU (icu)", "table": "procedureevents",
     "desc": "Procedures performed in the ICU (e.g. mechanical ventilation onset).",
     "cols": [
         ("stay_id",      "INT",       "FK", "icustays",  "ICU stay identifier.",        "30000026"),
         ("subject_id",   "INT",       "FK", "patients",  "Patient identifier.",         "10000032"),
         ("itemid",       "INT",       "FK", "d_items",   "Procedure item identifier.",  "225792"),
         ("charttime",    "TIMESTAMP", "",   "",          "Procedure start time.",       "2180-05-07 10:00"),
         ("value",        "DOUBLE",    "",   "",          "Duration or dose.",           "6.0"),
         ("valueuom",     "VARCHAR",   "",   "",          "Unit.",                       "hours"),
         ("caregiver_id", "INT",       "FK", "caregiver", "De-identified caregiver ID.", "22301"),
     ]},
    {"module": "ICU (icu)", "table": "datetimeevents",
     "desc": "Datetime-valued observations charted in the ICU.",
     "cols": [
         ("stay_id",      "INT",       "FK", "icustays",  "ICU stay identifier.",        "30000026"),
         ("subject_id",   "INT",       "FK", "patients",  "Patient identifier.",         "10000032"),
         ("itemid",       "INT",       "FK", "d_items",   "Datetime item identifier.",   "225775"),
         ("charttime",    "TIMESTAMP", "",   "",          "Chart time.",                 "2180-05-07 00:00"),
         ("value",        "TIMESTAMP", "",   "",          "Observed datetime value.",    "2180-05-06 22:40"),
         ("valueuom",     "VARCHAR",   "",   "",          "Unit.",                       "None"),
         ("caregiver_id", "INT",       "FK", "caregiver", "De-identified caregiver ID.", "22301"),
     ]},
    {"module": "ICU (icu)", "table": "ingredientevents",
     "desc": "Nutritional ingredient inputs (calories, protein, etc.) in the ICU.",
     "cols": [
         ("stay_id",    "INT",       "FK", "icustays", "ICU stay identifier.",   "30000026"),
         ("subject_id", "INT",       "FK", "patients", "Patient identifier.",    "10000032"),
         ("itemid",     "INT",       "FK", "d_items",  "Ingredient item ID.",    "220949"),
         ("charttime",  "TIMESTAMP", "",   "",         "Time of ingredient.",    "2180-05-07 06:00"),
         ("amount",     "DOUBLE",    "",   "",         "Amount of ingredient.",  "500.0"),
         ("amountuom",  "VARCHAR",   "",   "",         "Amount unit.",           "ml"),
     ]},
    {"module": "ICU (icu)", "table": "caregiver",
     "desc": "De-identified caregiver (nurse/doctor) identifiers and roles.",
     "cols": [
         ("caregiver_id", "INT",     "PK", "", "Unique caregiver identifier.", "22301"),
         ("label",        "VARCHAR", "",   "", "Caregiver role label.",        "Registered Nurse"),
     ]},

    # ══ DERIVED ══
    {"module": "Derived (derived)", "table": "sepsis3",
     "desc": "Binary Sepsis-3 indicator per ICU stay (SOFA ≥ 2 + suspected infection).",
     "cols": [
         ("stay_id",    "INT",     "PK", "icustays", "ICU stay identifier.",               "30000026"),
         ("subject_id", "INT",     "FK", "patients", "Patient identifier.",                "10000032"),
         ("sepsis3",    "BOOLEAN", "",   "",         "True if Sepsis-3 criteria are met.", "True"),
     ]},
    {"module": "Derived (derived)", "table": "sofa",
     "desc": "Sequential Organ Failure Assessment score per ICU day (0-24).",
     "cols": [
         ("stay_id",              "INT",  "PK", "icustays", "ICU stay identifier.",             "30000026"),
         ("subject_id",           "INT",  "FK", "patients", "Patient identifier.",              "10000032"),
         ("chartdate",            "DATE", "",   "",         "Date of SOFA scoring.",            "2180-05-07"),
         ("sofa_score",           "INT",  "",   "",         "Composite SOFA score (0-24).",     "8"),
         ("respiration_score",    "INT",  "",   "",         "Respiratory component (0-4).",     "2"),
         ("coagulation_score",    "INT",  "",   "",         "Coagulation component (0-4).",     "1"),
         ("liver_score",          "INT",  "",   "",         "Liver component (0-4).",           "0"),
         ("cardiovascular_score", "INT",  "",   "",         "Cardiovascular component (0-4).",  "3"),
         ("cns_score",            "INT",  "",   "",         "CNS component (GCS-based, 0-4).",  "1"),
         ("renal_score",          "INT",  "",   "",         "Renal component (0-4).",           "1"),
     ]},
    {"module": "Derived (derived)", "table": "sapsii",
     "desc": "SAPS II score: Simplified Acute Physiology Score II.",
     "cols": [
         ("stay_id",      "INT",  "PK", "icustays", "ICU stay identifier.",                    "30000026"),
         ("subject_id",   "INT",  "FK", "patients", "Patient identifier.",                     "10000032"),
         ("chartdate",    "DATE", "",   "",         "Scoring date.",                           "2180-05-07"),
         ("sapsii_score", "INT",  "",   "",         "SAPS II score (0-163); higher = worse.",  "42"),
     ]},
    {"module": "Derived (derived)", "table": "apsiii",
     "desc": "APS III score: Acute Physiology Score III.",
     "cols": [
         ("stay_id",      "INT", "PK", "icustays", "ICU stay identifier.", "30000026"),
         ("subject_id",   "INT", "FK", "patients", "Patient identifier.",  "10000032"),
         ("apsiii_score", "INT", "",   "",         "APS III score.",       "55"),
     ]},
    {"module": "Derived (derived)", "table": "oasis",
     "desc": "OASIS score: Oxford Acute Severity of Illness Score.",
     "cols": [
         ("stay_id",    "INT", "PK", "icustays", "ICU stay identifier.", "30000026"),
         ("subject_id", "INT", "FK", "patients", "Patient identifier.",  "10000032"),
         ("oasis_score","INT", "",   "",         "OASIS score.",         "30"),
     ]},
    {"module": "Derived (derived)", "table": "lods",
     "desc": "LODS score: Logistic Organ Dysfunction System.",
     "cols": [
         ("stay_id",    "INT", "PK", "icustays", "ICU stay identifier.", "30000026"),
         ("subject_id", "INT", "FK", "patients", "Patient identifier.",  "10000032"),
         ("lods_score", "INT", "",   "",         "LODS score.",          "4"),
     ]},
    {"module": "Derived (derived)", "table": "charlson",
     "desc": "Charlson Comorbidity Index per ICU stay.",
     "cols": [
         ("stay_id",        "INT", "PK", "icustays", "ICU stay identifier.",       "30000026"),
         ("subject_id",     "INT", "FK", "patients", "Patient identifier.",        "10000032"),
         ("charlson_score", "INT", "",   "",         "Charlson score (0→37+).",   "4"),
     ]},
    {"module": "Derived (derived)", "table": "ventilation",
     "desc": "Mechanical ventilation periods per ICU stay.",
     "cols": [
         ("stay_id",            "INT",       "FK", "icustays", "ICU stay identifier.",     "30000026"),
         ("subject_id",         "INT",       "FK", "patients", "Patient identifier.",      "10000032"),
         ("vent_start",         "TIMESTAMP", "",   "",         "Ventilation start time.",  "2180-05-07 10:00"),
         ("vent_end",           "TIMESTAMP", "",   "",         "Ventilation end time.",    "2180-05-10 08:30"),
         ("ventilation_status", "VARCHAR",   "",   "",         "Vent type/status.",        "InvasiveVent"),
     ]},
    {"module": "Derived (derived)", "table": "rrt",
     "desc": "Renal replacement therapy (dialysis) periods.",
     "cols": [
         ("stay_id",    "INT",       "FK", "icustays", "ICU stay identifier.", "30000026"),
         ("subject_id", "INT",       "FK", "patients", "Patient identifier.",  "10000032"),
         ("rrt_start",  "TIMESTAMP", "",   "",         "RRT start time.",      "2180-05-09 08:00"),
     ]},
    {"module": "Derived (derived)", "table": "kdigo_stages",
     "desc": "Acute kidney injury staging per KDIGO criteria.",
     "cols": [
         ("stay_id",     "INT",  "FK", "icustays", "ICU stay identifier.",              "30000026"),
         ("subject_id",  "INT",  "FK", "patients", "Patient identifier.",               "10000032"),
         ("chartdate",   "DATE", "",   "",         "Assessment date.",                  "2180-05-08"),
         ("kdigo_stage", "INT",  "",   "",         "AKI stage (1-3; null = no AKI).",  "2"),
     ]},
    {"module": "Derived (derived)", "table": "norepinephrine_equivalent_dose",
     "desc": "Vasopressor load expressed as norepinephrine equivalent dose.",
     "cols": [
         ("stay_id",             "INT",    "FK", "icustays", "ICU stay identifier.",             "30000026"),
         ("subject_id",          "INT",    "FK", "patients", "Patient identifier.",              "10000032"),
         ("chartdate",           "DATE",   "",   "",         "Dose calculation date.",           "2180-05-07"),
         ("norepinephrine_rate", "DOUBLE", "",   "",         "NE-equivalent dose (mcg/kg/min).", "0.08"),
     ]},
    {"module": "Derived (derived)", "table": "chemistry",
     "desc": "Chemistry lab values aggregated per ICU day.",
     "cols": [
         ("stay_id",    "INT",    "PK", "icustays", "ICU stay identifier.",         "30000026"),
         ("subject_id", "INT",    "FK", "patients", "Patient identifier.",          "10000032"),
         ("chartdate",  "DATE",   "",   "",         "Lab measurement date.",        "2180-05-07"),
         ("creatinine", "DOUBLE", "",   "",         "Creatinine (mg/dL).",          "1.4"),
         ("bun",        "DOUBLE", "",   "",         "Blood urea nitrogen (mg/dL).", "22"),
         ("sodium",     "DOUBLE", "",   "",         "Sodium (mEq/L).",              "138"),
         ("potassium",  "DOUBLE", "",   "",         "Potassium (mEq/L).",           "4.0"),
         ("chloride",   "DOUBLE", "",   "",         "Chloride (mEq/L).",            "100"),
         ("bicarbonate","DOUBLE", "",   "",         "Bicarbonate (mEq/L).",         "22"),
         ("glucose",    "DOUBLE", "",   "",         "Glucose (mg/dL).",             "140"),
     ]},
    {"module": "Derived (derived)", "table": "complete_blood_count",
     "desc": "CBC lab values aggregated per ICU day.",
     "cols": [
         ("stay_id",    "INT",    "PK", "icustays", "ICU stay identifier.",        "30000026"),
         ("subject_id", "INT",    "FK", "patients", "Patient identifier.",         "10000032"),
         ("chartdate",  "DATE",   "",   "",         "Lab measurement date.",       "2180-05-07"),
         ("wbc",        "DOUBLE", "",   "",         "WBC count (×10³/µL).",       "15.2"),
         ("hemoglobin", "DOUBLE", "",   "",         "Hemoglobin (g/dL).",          "10.4"),
         ("hematocrit", "DOUBLE", "",   "",         "Hematocrit (%).",             "31.8"),
         ("platelet",   "DOUBLE", "",   "",         "Platelet count (×10³/µL).",  "180"),
     ]},
    {"module": "Derived (derived)", "table": "coagulation",
     "desc": "Coagulation lab values aggregated per ICU day.",
     "cols": [
         ("stay_id",    "INT",    "PK", "icustays", "ICU stay identifier.",            "30000026"),
         ("subject_id", "INT",    "FK", "patients", "Patient identifier.",             "10000032"),
         ("chartdate",  "DATE",   "",   "",         "Lab measurement date.",           "2180-05-07"),
         ("pt",         "DOUBLE", "",   "",         "Prothrombin time (seconds).",     "14.2"),
         ("inr",        "DOUBLE", "",   "",         "International normalised ratio.", "1.2"),
         ("ptt",        "DOUBLE", "",   "",         "Partial thromboplastin time (s).","38.7"),
     ]},
    {"module": "Derived (derived)", "table": "inflammation",
     "desc": "Inflammation markers per ICU day.",
     "cols": [
         ("stay_id",    "INT",    "PK", "icustays", "ICU stay identifier.",       "30000026"),
         ("subject_id", "INT",    "FK", "patients", "Patient identifier.",        "10000032"),
         ("chartdate",  "DATE",   "",   "",         "Lab measurement date.",      "2180-05-07"),
         ("crp",        "DOUBLE", "",   "",         "C-reactive protein (mg/L).", "120.5"),
     ]},
    {"module": "Derived (derived)", "table": "first_day_vitalsign",
     "desc": "Min/max vital signs on ICU admission day 1.",
     "cols": [
         ("stay_id",       "INT",    "PK", "icustays","ICU stay identifier.",         "30000026"),
         ("subject_id",    "INT",    "FK", "patients","Patient identifier.",          "10000032"),
         ("heartrate_min", "DOUBLE", "",   "",        "Min heart rate (bpm).",        "72"),
         ("heartrate_max", "DOUBLE", "",   "",        "Max heart rate (bpm).",        "112"),
         ("sysbp_min",     "DOUBLE", "",   "",        "Min systolic BP (mmHg).",      "88"),
         ("sysbp_max",     "DOUBLE", "",   "",        "Max systolic BP (mmHg).",      "145"),
         ("tempc_min",     "DOUBLE", "",   "",        "Min temperature (°C).",        "36.4"),
         ("tempc_max",     "DOUBLE", "",   "",        "Max temperature (°C).",        "38.7"),
         ("spo2_min",      "DOUBLE", "",   "",        "Min SpO₂ (%).",                "90"),
         ("spo2_max",      "DOUBLE", "",   "",        "Max SpO₂ (%).",                "99"),
         ("resprate_min",  "DOUBLE", "",   "",        "Min respiratory rate (bpm).",  "14"),
         ("resprate_max",  "DOUBLE", "",   "",        "Max respiratory rate (bpm).",  "28"),
     ]},
    {"module": "Derived (derived)", "table": "first_day_lab",
     "desc": "Key lab values observed on ICU day 1.",
     "cols": [
         ("stay_id",        "INT",    "PK","icustays","ICU stay identifier.",          "30000026"),
         ("subject_id",     "INT",    "FK","patients","Patient identifier.",           "10000032"),
         ("creatinine_min", "DOUBLE", "",  "",        "Min creatinine (mg/dL).",      "0.9"),
         ("creatinine_max", "DOUBLE", "",  "",        "Max creatinine (mg/dL).",      "1.4"),
         ("sodium_min",     "DOUBLE", "",  "",        "Min sodium (mEq/L).",          "132"),
         ("sodium_max",     "DOUBLE", "",  "",        "Max sodium (mEq/L).",          "140"),
         ("wbc_min",        "DOUBLE", "",  "",        "Min WBC (×10³/µL).",          "8.0"),
         ("wbc_max",        "DOUBLE", "",  "",        "Max WBC (×10³/µL).",          "18.4"),
         ("lactate_min",    "DOUBLE", "",  "",        "Min lactate (mmol/L).",        "1.1"),
         ("lactate_max",    "DOUBLE", "",  "",        "Max lactate (mmol/L).",        "3.4"),
     ]},
    {"module": "Derived (derived)", "table": "first_day_gcs",
     "desc": "Glasgow Coma Scale score on ICU day 1.",
     "cols": [
         ("stay_id",        "INT","PK","icustays","ICU stay identifier.",       "30000026"),
         ("subject_id",     "INT","FK","patients","Patient identifier.",        "10000032"),
         ("gcs_min",        "INT","",  "",        "Min GCS total (3-15).",     "12"),
         ("gcs_eyes_min",   "INT","",  "",        "Min eye response (1-4).",   "3"),
         ("gcs_verbal_min", "INT","",  "",        "Min verbal response (1-5).","3"),
         ("gcs_motor_min",  "INT","",  "",        "Min motor response (1-6).", "6"),
     ]},
    {"module": "Derived (derived)", "table": "first_day_urine_output",
     "desc": "Total urine output on ICU day 1.",
     "cols": [
         ("stay_id",      "INT",   "PK","icustays","ICU stay identifier.",           "30000026"),
         ("subject_id",   "INT",   "FK","patients","Patient identifier.",            "10000032"),
         ("urine_output", "DOUBLE","",  "",        "Total urine output day 1 (mL).","1250.0"),
     ]},
]

# ─── V13b and Vital Signs (unchanged) ─────────────────────────────
vital_cols = ["signal_name","mimic_source_table","itemid","unit","example_values","notes"]
vital_rows = [
    ["heart_rate",           "chartevents", "220045",           "bpm",     "88 / 112 / 72",    "Heart rate from bedside monitor. Forward-filled at 1 h resolution."],
    ["resprate",             "chartevents", "220210, 224422",   "bpm",     "22 / 18 / 28",     "Respiratory rate from bedside monitor or manual count."],
    ["spo2",                 "chartevents", "220277",           "%",       "94 / 88 / 99",     "Peripheral oxygen saturation (pulse oximetry)."],
    ["temp_c",               "chartevents", "223762, 226329",   "°C",      "38.7 / 39.4 / 36.2","Body temperature. Fahrenheit values converted to Celsius."],
    ["sbp",                  "chartevents", "220179, 224167",   "mmHg",    "102 / 72 / 160",   "Systolic blood pressure (non-invasive or arterial line)."],
    ["dbp",                  "chartevents", "220180, 224643",   "mmHg",    "64 / 48 / 95",     "Diastolic blood pressure (non-invasive or arterial line)."],
    ["map",                  "derived",     "computed",         "mmHg",    "77 / 56 / 117",    "Mean arterial pressure = (SBP + 2×DBP) / 3. Derived in pipeline."],
    ["weight_kg",            "chartevents", "226512",           "kg",      "85 / 112 / 52",    "Admission weight. Used for dose normalisation and derived features."],
    ["age",                  "patients",    "anchor_age + offset","years", "52 / 71 / 43",     "Age at ICU admission. Derived from anchor_age and admission year."],
    ["time_since_admission", "derived",     "charttime − intime","hours", "0.0 / 12.5 / 48.3","Hours elapsed since ICU intime at each measurement row."],
]

ml_cols = ["group","feature_name","engineering_type","formula_origin","interpretation"]
ml_rows = [
    ["Base vital",       "heart_rate",                       "Raw",              "chartevents 220045",                                           "Heart rate (bpm)"],
    ["Base vital",       "resprate",                         "Raw",              "chartevents 220210 / 224422",                                  "Respiratory rate (bpm)"],
    ["Base vital",       "spo2",                             "Raw",              "chartevents 220277",                                           "SpO₂ (%)"],
    ["Base vital",       "temp_c",                           "Raw",              "chartevents 223762 / 226329",                                  "Body temperature (°C)"],
    ["Base vital",       "sbp",                              "Raw",              "chartevents 220179 / 224167",                                  "Systolic blood pressure (mmHg)"],
    ["Base vital",       "dbp",                              "Raw",              "chartevents 220180 / 224643",                                  "Diastolic blood pressure (mmHg)"],
    ["Base vital",       "weight_kg",                        "Raw",              "chartevents 226512",                                           "Patient weight (kg)"],
    ["Base vital",       "age",                              "Raw",              "patients.anchor_age + year offset",                            "Age at ICU admission (years)"],
    ["Base vital",       "time_since_admission",             "Derived",          "charttime − icustays.intime",                                  "Hours elapsed since ICU admission"],
    ["Composite",        "shock_index",                      "Composite",        "heart_rate / sbp",                                             "Shock index — elevated = cardiovascular compromise"],
    ["Composite",        "map",                              "Composite",        "(sbp + 2×dbp) / 3",                                            "Mean arterial pressure (mmHg)"],
    ["Composite",        "partial_qsofa",                    "Composite",        "(sbp < 100) + (resprate ≥ 20)",                                "Partial qSOFA score (0-2)"],
    ["Composite",        "news_score",                       "Composite",        "Weighted sum: HR, RR, SpO₂, Temp, SBP",                        "National Early Warning Score (0-20)"],
    ["Composite",        "qsofa_delta",                      "Composite",        "partial_qsofa(t) − partial_qsofa(t-1)",                        "1-hour change in qSOFA"],
    ["Composite",        "news_delta",                       "Composite",        "news_score(t) − news_score(t-1)",                              "1-hour change in NEWS"],
    ["Expanding window", "exp_min_{vital}",                  "Expanding",        "min(vital) from ICU admission to t",                           "Running patient-specific minimum (×6 vitals)"],
    ["Expanding window", "exp_max_{vital}",                  "Expanding",        "max(vital) from ICU admission to t",                           "Running patient-specific maximum (×6 vitals)"],
    ["Expanding window", "exp_mean_{vital}",                 "Expanding",        "mean(vital) from ICU admission to t",                          "Running patient-specific mean (×6 vitals)"],
    ["Expanding window", "exp_std_{vital}",                  "Expanding",        "std(vital) from ICU admission to t",                           "Running patient-specific std. dev. (×6 vitals)"],
    ["Rolling / slope",  "{vital}_sd_4h",                   "Rolling 4 h",      "std(vital) over preceding 4 hourly rows",                      "Short-term variability (×6 vitals)"],
    ["Rolling / slope",  "{vital}_ewma_3h",                 "EWMA 3 h",         "Exponentially weighted mean, span=3",                          "Smoothed trend (×6 vitals)"],
    ["Rolling / slope",  "slope_4h_{vital}",                "Slope 4 h",        "vital(t) − vital(t-4)",                                        "4-hour slope proxy (×6 vitals)"],
    ["Lag kinematics",   "lag_diff_1h_{vital}",             "Lag difference",   "vital(t) − vital(t-1)",                                        "1-hour first-order change (×6 vitals)"],
    ["Lag kinematics",   "lag_ratio_1h_{vital}",            "Lag ratio",        "vital(t) / (vital(t-1) + ε)",                                  "1-hour relative change (×6 vitals)"],
    ["Interaction",      "ventilation_perfusion_proxy",      "Interaction",      "resprate × (100 − spo2)",                                      "V/Q mismatch proxy: rises when RR↑ and SpO₂↓"],
    ["Interaction",      "tachycardia_excess",               "Interaction",      "heart_rate − 10 × (temp_c − 37)",                              "HR above fever-expected threshold"],
    ["Interaction",      "perfusion_adequacy",               "Interaction",      "heart_rate / ((sbp + ε) × (spo2/100 + ε))",                   "Perfusion adequacy index"],
    ["Interaction",      "cardiorespiratory_coupling_4h",    "Rolling corr 4 h", "Pearson(heart_rate, resprate) over 4 h window",                "HR-RR coupling — decouples under physiological stress"],
    ["Acceleration",     "accel_{vital}",                    "2nd derivative",   "Δvital(t) − Δvital(t-1)",                                      "Vital sign acceleration / deceleration (×6 vitals)"],
    ["CUSUM",            "cusum_pos_{vital}",                "CUSUM",            "Cumulative sum of positive deviations from running mean",       "Accumulated upward drift above patient baseline (×6 vitals)"],
    ["CUSUM",            "cusum_neg_{vital}",                "CUSUM",            "Cumulative sum of |negative deviations| from running mean",     "Accumulated downward drift below patient baseline (×6 vitals)"],
    ["Composite (V13)",  "pulse_pressure",                   "Composite",        "sbp − dbp",                                                    "Pulse pressure (mmHg) — reflects stroke volume"],
    ["Composite (V13)",  "rpp",                              "Composite",        "heart_rate × sbp",                                             "Rate-pressure product — proxy for cardiac O₂ demand"],
    ["Composite (V13)",  "msi",                              "Composite",        "heart_rate / MAP",                                             "Modified shock index using MAP as denominator"],
    ["Composite (V13)",  "resp_distress",                    "Composite",        "resprate × (100 − spo2)",                                      "Respiratory distress index"],
    ["Composite (V13)",  "fever_tachycardia",                "Composite",        "heart_rate × max(temp_c − 37, 0)",                             "Joint fever-tachycardia severity signal"],
    ["Meta (V13b)",      "meta_is_sepsis_6h",                "NOSE OOF",         "RF ensemble on 6-h sepsis target (OOF probability)",           "Predicted probability of sepsis onset within 6 h"],
    ["Meta (V13b)",      "meta_is_sepsis_12h",               "NOSE OOF",         "RF ensemble on 12-h sepsis target (OOF probability)",          "Predicted probability of sepsis onset within 12 h"],
    ["Meta (V13b)",      "meta_target_resp_6h",              "NOSE OOF",         "Aetiology stream: respiratory, 6 h horizon",                   "OOF probability — respiratory-source sepsis at 6 h"],
    ["Meta (V13b)",      "meta_target_resp_12h",             "NOSE OOF",         "Aetiology stream: respiratory, 12 h horizon",                  "OOF probability — respiratory-source sepsis at 12 h"],
    ["Meta (V13b)",      "meta_target_uri_6h",               "NOSE OOF",         "Aetiology stream: urinary, 6 h horizon",                       "OOF probability — urinary-source sepsis at 6 h"],
    ["Meta (V13b)",      "meta_target_uri_12h",              "NOSE OOF",         "Aetiology stream: urinary, 12 h horizon",                      "OOF probability — urinary-source sepsis at 12 h"],
    ["Meta (V13b)",      "meta_target_other_6h",             "NOSE OOF",         "Aetiology stream: other/systemic, 6 h horizon",                "OOF probability — other-source sepsis at 6 h"],
    ["Meta (V13b)",      "meta_target_other_12h",            "NOSE OOF",         "Aetiology stream: other/systemic, 12 h horizon",               "OOF probability — other-source sepsis at 12 h"],
    ["Meta (V13b)",      "var_6h",                           "Ensemble variance","var(meta_*_6h across 4 streams)",                              "Disagreement between 6-h stream predictions"],
    ["Meta (V13b)",      "var_12h",                          "Ensemble variance","var(meta_*_12h across 4 streams)",                             "Disagreement between 12-h stream predictions"],
]

# ─── Build Workbook ───────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── V13b ML FEATURES SHEETS (First 2 sheets) ──
# 1. Explanation
mf_ex = wb.create_sheet("V13b ML Features expl")
mf_ex.freeze_panes = "A3"
write_section_header(mf_ex, 1, "V13b Model — Feature Explanations", len(ml_cols))
write_col_headers(mf_ex, 2, ml_cols)
for ri, row in enumerate(ml_rows, 3):
    write_data_row(mf_ex, ri, row, alt=(ri%2==0))
auto_col_widths(mf_ex, ml_cols, ml_rows, min_w=14, max_w=70)

# 2. Data Snippet
mf_da = wb.create_sheet("V13b ML Features")
mf_da.freeze_panes = "A3"

# True list of 96 engineered features + 10 meta-learner features
exact_features = [
    'age', 'weight_kg', 'time_since',
    'heart_rate', 'resprate', 'spo2', 'temp_c', 'sbp', 'dbp',
    'map', 'shock_index', 'partial_qsofa', 'news_score', 'qsofa_delta', 'news_delta',
    'pulse_pressure', 'rpp', 'msi', 'resp_distress', 'fever_tachycardia',
    'ventilation_perfusion_proxy', 'tachycardia_excess', 'perfusion_adequacy', 'cardiorespiratory_coupling_4h'
]
vitals = ['heart_rate', 'resprate', 'spo2', 'temp_c', 'sbp', 'dbp']
for v in vitals:
    exact_features.extend([
        f'exp_min_{v}', f'exp_max_{v}', f'exp_mean_{v}', f'exp_std_{v}',
        f'{v}_sd_4h', f'{v}_ewma_3h', f'slope_4h_{v}',
        f'lag_diff_1h_{v}', f'lag_ratio_1h_{v}',
        f'accel_{v}', f'cusum_pos_{v}', f'cusum_neg_{v}'
    ])

meta_features = [
    'meta_is_sepsis_6h', 'meta_is_sepsis_12h',
    'meta_target_resp_6h', 'meta_target_resp_12h',
    'meta_target_uri_6h', 'meta_target_uri_12h',
    'meta_target_other_6h', 'meta_target_other_12h',
    'var_6h', 'var_12h'
]
exact_features.extend(meta_features)

ml_feature_names = ["stay_id", "charttime"] + exact_features
write_section_header(mf_da, 1, "V13b Model — Features Data Snippet", len(ml_feature_names))
write_col_headers(mf_da, 2, ml_feature_names)

# Mock some data for 10 rows for all 106+ features
mf_mock_data = []
random.seed(42)  # EDIT: random seed for mock feature-snippet rows
start_stay_id = 30000026

for i in range(10):
    current_stay = start_stay_id + i  # Unique stay_id for each row
    row = [current_stay, f"2180-05-07 {i:02d}:00"]

    for f_name in exact_features:
        if "time_since" in f_name:
            val = i % 3
        elif "heart_rate" in f_name or "resprate" in f_name:
            val = round(random.uniform(70, 110), 1) if "rate" in f_name else round(random.uniform(14, 25), 1)
        elif "spo2" in f_name:
            val = random.randint(90, 100)
        elif "sbp" in f_name or "dbp" in f_name or "map" in f_name or "pulse_pressure" in f_name:
            val = random.randint(55, 145)
        elif "temp" in f_name:
            val = round(36.5 + random.uniform(0, 2), 1)
        elif "weight" in f_name:
            val = 85.2
        elif "age" in f_name:
            val = 65
        elif "meta" in f_name or "var" in f_name:
            val = round(random.uniform(0, 1), 4)
        elif "cusum" in f_name:
            val = round(random.uniform(0, 10), 2)
        else:
            val = round(random.uniform(-2.0, 5.0), 3)
        row.append(val)

    mf_mock_data.append(row)

for ri, row in enumerate(mf_mock_data, 3):
    write_data_row(mf_da, ri, row, alt=(ri%2==0))
auto_col_widths(mf_da, ml_feature_names, mf_mock_data, max_w=20)

# ── INDEX SHEET ──
idx = wb.create_sheet("INDEX")
idx.freeze_panes = "A3"
idx.column_dimensions["A"].width = 5
idx.column_dimensions["B"].width = 30
idx.column_dimensions["C"].width = 30
idx.column_dimensions["D"].width = 15
idx.column_dimensions["E"].width = 65

write_section_header(idx, 1, "MIMIC-IV v3.1 — Schema Index", 5)
write_col_headers(idx, 2, ["#", "Schema Definition Sheet", "Data Snippet Sheet", "Module", "Description"])

all_entries = [
    ("V13b ML Features expl", "V13b ML Features", "Special", len(ml_rows), "All features (raw + engineered) fed to the V13b machine learning model")
]
for t in schema:
    all_entries.append((f"{t['table'][:25]} expl", t["table"], t["module"], len(t["cols"]), t["desc"]))
all_entries.append(("Vital Signs Used", "-", "Special", len(vital_rows), "Raw vital sign signals extracted from MIMIC-IV used as pipeline inputs"))

for i, (expl_name, data_name, mod, nrows, desc) in enumerate(all_entries, 1):
    write_data_row(idx, i+2, [i, expl_name, data_name, mod, desc], ["center", "left", "left", "left", "left"], alt=(i%2==0))

# ── TABLE SHEETS ──
EXPL_COLS = ["#", "Column Name", "Data Type", "Key", "FK → Table", "Description", "Sample Value"]
EXPL_WID = [5, 28, 18, 5, 18, 55, 16]

for t in schema:
    raw_name = t["table"]
    expl_sname = (raw_name[:25] + " expl")
    data_sname = raw_name[:31]

    # 1. Explanation Sheet
    ws_ex = wb.create_sheet(expl_sname)
    ws_ex.freeze_panes = "A4"
    for ci, w in enumerate(EXPL_WID, 1):
        ws_ex.column_dimensions[get_column_letter(ci)].width = w

    write_section_header(ws_ex, 1, f"Schema Definition: {raw_name}  ({t['module']})", len(EXPL_COLS))
    ws_ex.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(EXPL_COLS))
    s = ws_ex.cell(row=2, column=1, value=t["desc"])
    s.font = f(italic=True, size=9, color="444444")
    s.fill = solid("EBEBEB")
    s.alignment = al("left")
    s.border = B
    ws_ex.row_dimensions[2].height = 16
    write_col_headers(ws_ex, 3, EXPL_COLS)

    for ri, col in enumerate(t["cols"], 4):
        name, dtype, key, fk, desc, sample = col
        write_data_row(ws_ex, ri,
                       [ri-3, name, dtype, key, fk, desc, str(sample)],
                       ["center","left","center","center","left","left","center"],
                       alt=(ri % 2 == 0))

    # 2. Data Snippet Sheet
    ws_da = wb.create_sheet(data_sname)
    ws_da.freeze_panes = "A3"
    col_names = [c[0] for c in t["cols"]]

    write_section_header(ws_da, 1, f"Data Snippet: {raw_name}", len(col_names))
    write_col_headers(ws_da, 2, col_names)

    mock_data = generate_mock_rows(t["cols"], n=10)
    for ri, row in enumerate(mock_data, 3):
        write_data_row(ws_da, ri, row, alt=(ri%2==0))

    auto_col_widths(ws_da, col_names, mock_data)

# ── VITAL SIGNS SHEET ──
vs = wb.create_sheet("Vital Signs Used")
vs.freeze_panes = "A3"
write_section_header(vs, 1, "Vital Signs Used — Raw Signals Extracted from MIMIC-IV", len(vital_cols))
write_col_headers(vs, 2, vital_cols)
for ri, row in enumerate(vital_rows, 3):
    write_data_row(vs, ri, row, alt=(ri%2==0))
auto_col_widths(vs, vital_cols, vital_rows, min_w=14, max_w=60)

# ─── Save ─────────────────────────────────────────────────────────
out = "/PATH/TO/PROJECT/technical/MIMIC_IV_Schema.xlsx"  # EDIT: output .xlsx workbook path
wb.save(out)
print(f"✅  Saved: {out}   ({len(wb.sheetnames)} sheets)")
