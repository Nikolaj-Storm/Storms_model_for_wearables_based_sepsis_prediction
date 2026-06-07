# Copyright (c) 2026 Nikolaj Storm Petersen. Licensed under CC BY-NC 4.0.
# Non-commercial use only. If you use or adapt this code, please cite the author.
# See LICENSE and CITATION.cff  |  https://creativecommons.org/licenses/by-nc/4.0/

# ============================================================================
#  run_sirs_proxy.py
#  Stage: 5 - Comparators
#
#  PURPOSE
#    Compute a SIRS-based sepsis-screen proxy (temperature, heart rate,
#    respiratory rate, optional WBC) per row, flag rows meeting the protocol
#    2-or-more criteria trigger, sweep cutoffs, and evaluate against the 6-hour
#    and 12-hour Sepsis-3 labels. Runs at a chosen sampling resolution and
#    appends operating points to the Excel workbook.
#
#  INPUTS
#    /PATH/TO/PROJECT/technical/Data/v3_dataset_<res>_test_full.parquet
#      where <res> is 1h, 15m, or 4h (chosen by CLI arg)
#  OUTPUTS
#    /PATH/TO/PROJECT/technical/Data/v3_dataset_<res>_test_SIRS.parquet
#    /PATH/TO/PROJECT/all results updated.xlsx  (rows appended)
#    Also prints sweep tables to the console.
#
#  USER-EDITABLE SETTINGS  (grep the body for the tag  EDIT:  to find each)
#    CLI arg            -  resolution key, one of 1h | 15m | 4h
#    RESOLUTION_CONFIG  -  per-resolution input/output parquet paths and label
#    EXCEL_PATH         -  results workbook the operating points are appended to
#    SIRS criteria cutoffs  -  temp > 38 or < 36, HR > 90, RR > 20, WBC bounds
#                          in compute_sirs; adjust to your screen definition
#    SIRS report cutoffs  -  the [1, 2, 3, 4] sweep with 2 as protocol trigger
#    target columns     -  is_sepsis_6h and is_sepsis_12h label columns
#
#  REQUIRES: numpy, pandas, openpyxl
# ============================================================================
"""
SIRS-based sepsis-screen proxy.

Faithful implementation of the sepsis-suspicion logic in the Danish TOKS
protocol. Per the protocol, "Sepsis = 2 SIRS + infektion." This proxy computes
the three SIRS criteria available from our cohort (temperature, heart rate,
respiratory rate) and flags rows where 2 or more are positive, treating that
flag as a sepsis-suspicion event.

The fourth SIRS criterion (white blood cell count) is not available in the
cohort and is acknowledged as a documented limitation. The "infection
suspicion" element of the protocol's definition is also unimplementable
retrospectively, since it requires real-time clinical judgment.

Run:
    python3 run_sirs_proxy.py 1h
    python3 run_sirs_proxy.py 15m
    python3 run_sirs_proxy.py 4h
"""

import sys
from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl

# EDIT: results workbook path
EXCEL_PATH = Path("/PATH/TO/PROJECT/all results updated.xlsx")
# EDIT: per-resolution input/output parquet paths
RESOLUTION_CONFIG = {
    "1h": {
        "input": Path("/PATH/TO/PROJECT/technical/Data/v3_dataset_1h_test_full.parquet"),
        "output": Path("/PATH/TO/PROJECT/technical/Data/v3_dataset_1h_test_SIRS.parquet"),
        "label": "1-hour",
    },
    "15m": {
        "input": Path("/PATH/TO/PROJECT/technical/Data/v3_dataset_15m_test_full.parquet"),
        "output": Path("/PATH/TO/PROJECT/technical/Data/v3_dataset_15m_test_SIRS.parquet"),
        "label": "15-minute",
    },
    "4h": {
        "input": Path("/PATH/TO/PROJECT/technical/Data/v3_dataset_4h_test_full.parquet"),
        "output": Path("/PATH/TO/PROJECT/technical/Data/v3_dataset_4h_test_SIRS.parquet"),
        "label": "4-hour",
    },
}


def compute_sirs(df: pd.DataFrame) -> pd.DataFrame:
    """Compute the four SIRS criteria per row.

    SIRS-1, body temperature > 38 °C or < 36 °C
    SIRS-2, heart rate > 90 bpm
    SIRS-3, respiratory rate > 20 breaths/min
    SIRS-4, WBC ≥ 12 × 10⁹/L or ≤ 4 × 10⁹/L (band-form-neutrophil sub-criterion
            not modelled, requires differential count not in cohort)

    If a criterion's underlying signal is missing, it is treated as 0 (criterion
    does not trigger), conservative for the proxy.
    """
    temp = df["temp_c"].to_numpy()
    hr = df["heart_rate"].to_numpy()
    rr = df["resprate"].to_numpy()

    # EDIT: SIRS temperature criterion cutoffs (> 38 °C or < 36 °C)
    crit_temp = ((temp > 38) | (temp < 36)).astype(int)
    crit_temp = np.where(np.isnan(temp), 0, crit_temp)

    # EDIT: SIRS heart-rate criterion cutoff (> 90 bpm)
    crit_hr = (hr > 90).astype(int)
    crit_hr = np.where(np.isnan(hr), 0, crit_hr)

    # EDIT: SIRS respiratory-rate criterion cutoff (> 20 breaths/min)
    crit_rr = (rr > 20).astype(int)
    crit_rr = np.where(np.isnan(rr), 0, crit_rr)

    if "SIRS_wbc" in df.columns:
        crit_wbc = df["SIRS_wbc"].fillna(0).astype(int).to_numpy()
    else:
        crit_wbc = np.zeros(len(df), dtype=int)

    sirs_score = crit_temp + crit_hr + crit_rr + crit_wbc

    out = pd.DataFrame({
        "SIRS_temp": crit_temp,
        "SIRS_hr": crit_hr,
        "SIRS_rr": crit_rr,
        "SIRS_wbc": crit_wbc,
        "SIRS_score": sirs_score,
        "is_sepsis_6h": df["is_sepsis_6h"].fillna(0).astype(int).values,
        "is_sepsis_12h": df["is_sepsis_12h"].fillna(0).astype(int).values,
    }, index=df.index)
    return out


def metrics_at_cutoff(scores, labels, k):
    flagged = scores >= k
    tp = int(((flagged) & (labels == 1)).sum())
    fp = int(((flagged) & (labels == 0)).sum())
    fn = int(((~flagged) & (labels == 1)).sum())
    tn = int(((~flagged) & (labels == 0)).sum())
    pos = tp + fn
    neg = fp + tn
    total = pos + neg
    recall = tp / pos if pos else 0.0
    fpr = fp / neg if neg else 0.0
    fnp = fn / pos if pos else 0.0
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    f1_pos = (2 * tp) / (2 * tp + fp + fn) if (2 * tp + fp + fn) > 0 else 0.0
    f1_neg = (2 * tn) / (2 * tn + fn + fp) if (2 * tn + fn + fp) > 0 else 0.0
    f1_macro = (f1_pos + f1_neg) / 2
    if precision + recall > 0:
        f0_5 = (1.25 * precision * recall) / (0.25 * precision + recall)
        f2 = (5 * precision * recall) / (4 * precision + recall)
    else:
        f0_5 = 0.0
        f2 = 0.0
    return {
        "Precision (PPV)": precision, "F0.5": f0_5, "F1 Macro": f1_macro,
        "F2": f2, "FPR": fpr, "FNP": fnp,
        "TP": tp, "FN": fn, "FP": fp, "TN": tn,
        "Threshold": k, "Recall": recall, "F1 (pos class)": f1_pos,
        "TP %": 100 * tp / total if total else 0.0,
        "FN %": 100 * fn / total if total else 0.0,
        "FP %": 100 * fp / total if total else 0.0,
        "TN %": 100 * tn / total if total else 0.0,
    }


def auroc_auprc(scores, labels):
    cutoffs = list(range(int(scores.min()), int(scores.max()) + 2))
    rows = [metrics_at_cutoff(scores, labels, k) for k in cutoffs]
    fpr = np.array([r["FPR"] for r in rows])
    tpr = np.array([r["Recall"] for r in rows])
    prec = np.array([r["Precision (PPV)"] for r in rows])
    order = np.argsort(fpr)
    auroc = float(np.trapz(tpr[order], fpr[order]))
    order = np.argsort(tpr)
    auprc = float(np.trapz(prec[order], tpr[order]))
    return auroc, auprc


def evaluate(scores, labels, label_name):
    print(f"\n  Sweep against {label_name}")
    auroc, auprc = auroc_auprc(scores, labels)
    print(f"    AUROC {auroc:.4f}, AUPRC {auprc:.4f}")
    rows = []
    # EDIT: SIRS report cutoffs. 1+, 2+ (protocol sepsis trigger), 3+, 4+
    for k in [1, 2, 3, 4]:
        m = metrics_at_cutoff(scores, labels, k)
        m["AUPRC"] = auprc
        m["AUROC"] = auroc
        if k == 2:
            m["_variant_label"] = "SIRS >= 2 (protocol sepsis trigger)"
        else:
            m["_variant_label"] = f"SIRS >= {k}"
        rows.append(m)
        print(f"    SIRS ≥{k}: TP={m['TP']:5d}, FP={m['FP']:7d}, "
              f"recall={m['Recall']:.4f}, FPR={m['FPR']:.4f}, F1m={m['F1 Macro']:.4f}")
    return rows


EXCEL_COLS = ["Version", "Model", "Variant / Operating Point", "AUPRC",
              "Precision (PPV)", "AUROC", "F0.5", "F1 Macro", "F2", "FPR",
              "FNP", "TP", "FN", "FP", "TN", "Threshold", "Recall",
              "F1 (pos class)", "TP %", "FN %", "FP %", "TN %",
              "Train Acc", "Test Acc"]


def append_to_excel(rows, sheet_name, model_label):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[sheet_name]
    write_row = ws.max_row + 1
    for r in rows:
        data = {
            "Version": "Rule-Based Baseline",
            "Model": model_label,
            "Variant / Operating Point": r["_variant_label"],
            "AUPRC": r["AUPRC"], "Precision (PPV)": r["Precision (PPV)"],
            "AUROC": r["AUROC"], "F0.5": r["F0.5"], "F1 Macro": r["F1 Macro"],
            "F2": r["F2"], "FPR": r["FPR"], "FNP": r["FNP"],
            "TP": r["TP"], "FN": r["FN"], "FP": r["FP"], "TN": r["TN"],
            "Threshold": r["Threshold"], "Recall": r["Recall"],
            "F1 (pos class)": r["F1 (pos class)"],
            "TP %": r["TP %"], "FN %": r["FN %"], "FP %": r["FP %"], "TN %": r["TN %"],
            "Train Acc": "", "Test Acc": "",
        }
        for col_idx, col_name in enumerate(EXCEL_COLS, 1):
            ws.cell(row=write_row, column=col_idx, value=data.get(col_name, ""))
        write_row += 1
    wb.save(EXCEL_PATH)


def main():
    if len(sys.argv) < 2 or sys.argv[1] not in RESOLUTION_CONFIG:
        print("Usage: python3 run_sirs_proxy.py [1h|15m|4h]")
        sys.exit(1)
    res_key = sys.argv[1]
    cfg = RESOLUTION_CONFIG[res_key]

    print(f"\n=== SIRS-based sepsis-screen proxy at {cfg['label']} resolution ===")
    print(f"Reading {cfg['input']}...")
    df = pd.read_parquet(cfg["input"])
    print(f"  {df.shape[0]:,} rows, {df.index.get_level_values('stay_id').nunique():,} stays")
    print(f"  is_sepsis_6h prevalence:  {df['is_sepsis_6h'].mean():.4f}")
    print(f"  is_sepsis_12h prevalence: {df['is_sepsis_12h'].mean():.4f}")

    print(f"\nComputing SIRS criteria...")
    sirs = compute_sirs(df)
    print(f"  SIRS_temp positive rate: {sirs['SIRS_temp'].mean():.4f}")
    print(f"  SIRS_hr positive rate:   {sirs['SIRS_hr'].mean():.4f}")
    print(f"  SIRS_rr positive rate:   {sirs['SIRS_rr'].mean():.4f}")
    print(f"  SIRS_score >=2 rate:     {(sirs['SIRS_score'] >= 2).mean():.4f}")

    print(f"\nSaving SIRS scores to {cfg['output']}...")
    sirs.to_parquet(cfg["output"])

    rows_6h = evaluate(sirs["SIRS_score"].to_numpy(),
                       sirs["is_sepsis_6h"].to_numpy(),
                       "is_sepsis_6h")
    rows_12h = evaluate(sirs["SIRS_score"].to_numpy(),
                        sirs["is_sepsis_12h"].to_numpy(),
                        "is_sepsis_12h")

    model_label = f"SIRS-based sepsis screen ({cfg['label']})"
    print(f"\nAppending results to {EXCEL_PATH}...")
    append_to_excel(rows_6h, "6h Sepsis Prediction", model_label)
    append_to_excel(rows_12h, "12h Sepsis Prediction", model_label)
    print("Done.")


if __name__ == "__main__":
    main()
