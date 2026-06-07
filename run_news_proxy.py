# Copyright (c) 2026 Nikolaj Storm Petersen. Licensed under CC BY-NC 4.0.
# Non-commercial use only. If you use or adapt this code, please cite the author.
# See LICENSE and CITATION.cff  |  https://creativecommons.org/licenses/by-nc/4.0/

# ============================================================================
#  run_news_proxy.py
#  Stage: 5 - Comparators
#
#  PURPOSE
#    Score a Danish TOKS 2.1 (NEWS-variant) rule-based proxy on the 4-hour test
#    grid, sweep every integer cutoff, and evaluate against the 6-hour and
#    12-hour Sepsis-3 labels. Replicates the Nemati et al. (2018) NEWS setup
#    with Region Nordjylland thresholds. Appends operating points to the Excel
#    results workbook.
#
#  INPUTS
#    /PATH/TO/PROJECT/technical/Data/v3_dataset_4h_test_full.parquet
#  OUTPUTS
#    /PATH/TO/PROJECT/technical/Data/v3_dataset_4h_test_NEWS.parquet
#    /PATH/TO/PROJECT/all results updated.xlsx  (rows appended)
#    Also prints sweep tables to the console.
#
#  USER-EDITABLE SETTINGS  (grep the body for the tag  EDIT:  to find each)
#    INPUT             -  test parquet to score, set to your 4-hour test file
#    OUTPUT            -  parquet of per-row TOKS scores to write
#    EXCEL_PATH        -  results workbook the operating points are appended to
#    TOKS scoring thresholds  -  per-variable cutoffs in compute_news (RR, SpO2,
#                         SBP, HR, GCS, Temp); adjust to your EWS protocol
#    cutoffs_to_report -  clinical escalation cutoffs [1, 6, 8, 10] (Gron/Gul/
#                         Orange/Rod tiers)
#    target columns    -  is_sepsis_6h and is_sepsis_12h label columns
#
#  REQUIRES: numpy, pandas, openpyxl
# ============================================================================
"""
Danish TOKS 2.1 proxy on the cohort, replicating the Nemati et al. (2018) NEWS
evaluation setup with Danish-specific thresholds. Region Nordjylland TOKS 2.1
thresholds (Wellington-EWS-influenced NEWS variant), every-row scoring on the
4-hour grid, evaluated against `is_sepsis_6h` and `is_sepsis_12h`. No cadence
simulation, no red-flag overrides, no MAT criteria — just plain TOKS scoring.

Direct comparison reference: Nemati et al. (2018) reported NEWS AUROC 0.713 at
the 6-hour-before-onset horizon on a MIMIC-style ICU cohort with Sepsis-3
labels.
"""

import sys
from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl

# EDIT: input/output/workbook paths
INPUT = Path("/PATH/TO/PROJECT/technical/Data/v3_dataset_4h_test_full.parquet")
OUTPUT = Path("/PATH/TO/PROJECT/technical/Data/v3_dataset_4h_test_NEWS.parquet")
EXCEL_PATH = Path("/PATH/TO/PROJECT/all results updated.xlsx")


def compute_news(df: pd.DataFrame) -> pd.DataFrame:
    """Danish TOKS 2.1 scoring per Region Nordjylland (2025) protocol.

    Wellington-EWS-influenced variant of NEWS, with thresholds that shift by
    1 unit relative to NEWS2 in several variables. Aggregate score range
    0-20.
    """
    rr = df["resprate"].to_numpy()
    spo2 = df["spo2"].to_numpy()
    sbp = df["sbp"].to_numpy()
    hr = df["heart_rate"].to_numpy()
    gcs = df["GCS_total"].to_numpy()
    temp = df["temp_c"].to_numpy()
    on_o2 = df["on_O2"].to_numpy()

    # EDIT: RR scoring thresholds. ≤7→3, 8-11→1, 12-20→0, 21-24→2, ≥25→3
    pts_rr = np.where(rr <= 7, 3,
              np.where(rr <= 11, 1,
              np.where(rr <= 20, 0,
              np.where(rr <= 24, 2, 3))))
    pts_rr = np.where(np.isnan(rr), 0, pts_rr).astype(int)

    # EDIT: SpO2 scoring thresholds. ≥96→0, 94-95→1, 92-93→2, ≤91→3, +2 if on O2
    pts_spo2_base = np.where(spo2 >= 96, 0,
                     np.where(spo2 >= 94, 1,
                     np.where(spo2 >= 92, 2, 3)))
    pts_spo2_base = np.where(np.isnan(spo2), 0, pts_spo2_base)
    pts_spo2 = (pts_spo2_base + np.where(on_o2 == 1, 2, 0)).astype(int)
    pts_o2 = np.zeros(len(df), dtype=int)  # not separate, folded into SpO2

    # EDIT: SBP scoring thresholds. ≥220→3, 110-219→0, 100-109→1, 90-99→2, ≤89→3
    pts_sbp = np.where(sbp >= 220, 3,
               np.where(sbp >= 110, 0,
               np.where(sbp >= 100, 1,
               np.where(sbp >= 90, 2, 3))))
    pts_sbp = np.where(np.isnan(sbp), 0, pts_sbp).astype(int)

    # EDIT: HR scoring thresholds. ≥130→3, 110-129→2, 90-109→1, 50-89→0, 40-49→1, ≤39→3
    pts_hr = np.where(hr >= 110, np.where(hr >= 130, 3, 2),
              np.where(hr >= 90, 1,
              np.where(hr >= 50, 0,
              np.where(hr >= 40, 1, 3))))
    pts_hr = np.where(np.isnan(hr), 0, pts_hr).astype(int)

    # EDIT: Consciousness (GCS) scoring. 15→0, else→3
    pts_gcs = np.where(gcs == 15, 0, 3)
    pts_gcs = np.where(np.isnan(gcs), 0, pts_gcs).astype(int)

    # EDIT: Temperature scoring thresholds. ≥39→2, 38-38.9→1, 36-37.9→0, 35-35.9→1, ≤34.9→3
    pts_temp = np.where(temp >= 39, 2,
                np.where(temp >= 38, 1,
                np.where(temp >= 36, 0,
                np.where(temp >= 35, 1, 3))))
    pts_temp = np.where(np.isnan(temp), 0, pts_temp).astype(int)

    news_total = (pts_rr + pts_spo2 + pts_sbp + pts_hr + pts_gcs + pts_temp).astype(int)

    out = pd.DataFrame({
        "TOKS_resp": pts_rr, "TOKS_spo2": pts_spo2,
        "TOKS_sbp": pts_sbp, "TOKS_hr": pts_hr, "TOKS_gcs": pts_gcs,
        "TOKS_temp": pts_temp,
        "TOKS_total": news_total,
        "is_sepsis_6h": df["is_sepsis_6h"].fillna(0).astype(int).values,
        "is_sepsis_12h": df["is_sepsis_12h"].fillna(0).astype(int).values,
    }, index=df.index)
    return out


def metrics_at_cutoff(scores, labels, k):
    flagged = scores >= k
    tp = int(((flagged) & (labels == 1)).sum())
    fp = int(((flagged) & (labels == 0)).sum())
    fn = int(((~flagged) & (labels == 1)).sum())
    tn = int(((~flagged) & (labels == 0)).sum())
    pos = tp + fn
    neg = fp + tn
    total = pos + neg
    recall = tp / pos if pos else 0.0
    fpr = fp / neg if neg else 0.0
    fnp = fn / pos if pos else 0.0
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    f1_pos = (2 * tp) / (2 * tp + fp + fn) if (2 * tp + fp + fn) > 0 else 0.0
    f1_neg = (2 * tn) / (2 * tn + fn + fp) if (2 * tn + fn + fp) > 0 else 0.0
    f1_macro = (f1_pos + f1_neg) / 2
    if precision + recall > 0:
        f0_5 = (1.25 * precision * recall) / (0.25 * precision + recall)
        f2 = (5 * precision * recall) / (4 * precision + recall)
    else:
        f0_5 = 0.0
        f2 = 0.0
    return {
        "Precision (PPV)": precision, "F0.5": f0_5, "F1 Macro": f1_macro,
        "F2": f2, "FPR": fpr, "FNP": fnp,
        "TP": tp, "FN": fn, "FP": fp, "TN": tn,
        "Threshold": k, "Recall": recall, "F1 (pos class)": f1_pos,
        "TP %": 100 * tp / total if total else 0.0,
        "FN %": 100 * fn / total if total else 0.0,
        "FP %": 100 * fp / total if total else 0.0,
        "TN %": 100 * tn / total if total else 0.0,
    }


def auroc_auprc(scores, labels):
    cutoffs = list(range(int(scores.min()), int(scores.max()) + 2))
    rows = [metrics_at_cutoff(scores, labels, k) for k in cutoffs]
    fpr = np.array([r["FPR"] for r in rows])
    tpr = np.array([r["Recall"] for r in rows])
    prec = np.array([r["Precision (PPV)"] for r in rows])
    order = np.argsort(fpr)
    auroc = float(np.trapz(tpr[order], fpr[order]))
    order = np.argsort(tpr)
    auprc = float(np.trapz(prec[order], tpr[order]))
    return auroc, auprc


def evaluate(scores, labels, label_name):
    print(f"\n  Sweep against {label_name}")
    auroc, auprc = auroc_auprc(scores, labels)
    print(f"    AUROC {auroc:.4f}, AUPRC {auprc:.4f}")
    rows = []
    score_max = int(scores.max())
    f1m_best = (-1, -1.0)
    for k in range(0, score_max + 2):
        m = metrics_at_cutoff(scores, labels, k)
        if m["F1 Macro"] > f1m_best[1]:
            f1m_best = (k, m["F1 Macro"])
    f1m_cutoff = f1m_best[0]
    # EDIT: Danish TOKS 2.1 escalation cutoffs (Region Nordjylland 2025).
    cutoffs_to_report = [1, 6, 8, 10]
    if f1m_cutoff not in cutoffs_to_report:
        cutoffs_to_report.append(f1m_cutoff)
    cutoff_names = {1: "Threshold >= 1 (Grøn)",
                    6: "Threshold >= 6 (Gul)",
                    8: "Threshold >= 8 (Orange)",
                    10: "Threshold >= 10 (Rød)"}
    for k in cutoffs_to_report:
        m = metrics_at_cutoff(scores, labels, k)
        m["AUPRC"] = auprc
        m["AUROC"] = auroc
        if k == f1m_cutoff and k not in cutoff_names:
            m["_variant_label"] = f"Max F1-Macro (TOKS >= {k})"
        else:
            m["_variant_label"] = cutoff_names.get(k, f"TOKS >= {k}")
        rows.append(m)
        print(f"    TOKS ≥{k}: TP={m['TP']:5d}, FP={m['FP']:7d}, "
              f"recall={m['Recall']:.4f}, FPR={m['FPR']:.4f}, F1m={m['F1 Macro']:.4f}")
    return rows


EXCEL_COLS = ["Version", "Model", "Variant / Operating Point", "AUPRC",
              "Precision (PPV)", "AUROC", "F0.5", "F1 Macro", "F2", "FPR",
              "FNP", "TP", "FN", "FP", "TN", "Threshold", "Recall",
              "F1 (pos class)", "TP %", "FN %", "FP %", "TN %",
              "Train Acc", "Test Acc"]


def append_to_excel(rows, sheet_name, model_label):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[sheet_name]
    write_row = ws.max_row + 1
    for r in rows:
        data = {
            "Version": "Rule-Based Baseline",
            "Model": model_label,
            "Variant / Operating Point": r["_variant_label"],
            "AUPRC": r["AUPRC"], "Precision (PPV)": r["Precision (PPV)"],
            "AUROC": r["AUROC"], "F0.5": r["F0.5"], "F1 Macro": r["F1 Macro"],
            "F2": r["F2"], "FPR": r["FPR"], "FNP": r["FNP"],
            "TP": r["TP"], "FN": r["FN"], "FP": r["FP"], "TN": r["TN"],
            "Threshold": r["Threshold"], "Recall": r["Recall"],
            "F1 (pos class)": r["F1 (pos class)"],
            "TP %": r["TP %"], "FN %": r["FN %"], "FP %": r["FP %"], "TN %": r["TN %"],
            "Train Acc": "", "Test Acc": "",
        }
        for col_idx, col_name in enumerate(EXCEL_COLS, 1):
            ws.cell(row=write_row, column=col_idx, value=data.get(col_name, ""))
        write_row += 1
    wb.save(EXCEL_PATH)


def main():
    print(f"\n=== Danish TOKS 2.1 proxy at 4-hour resolution (Nemati 2018 replication) ===")
    print(f"Reading {INPUT}...")
    df = pd.read_parquet(INPUT)
    print(f"  {df.shape[0]:,} rows, {df.index.get_level_values('stay_id').nunique():,} stays")
    print(f"  is_sepsis_6h prevalence:  {df['is_sepsis_6h'].mean():.4f}")
    print(f"  is_sepsis_12h prevalence: {df['is_sepsis_12h'].mean():.4f}")

    news = compute_news(df)
    print(f"\n  TOKS_total range: {news['TOKS_total'].min()}..{news['TOKS_total'].max()}")
    print(f"  TOKS_total mean:  {news['TOKS_total'].mean():.2f}")
    print(f"  TOKS ≥3 rate:     {(news['TOKS_total'] >= 3).mean():.4f}")
    print(f"  TOKS ≥5 rate:     {(news['TOKS_total'] >= 5).mean():.4f}")
    print(f"  TOKS ≥7 rate:     {(news['TOKS_total'] >= 7).mean():.4f}")

    print(f"\nSaving TOKS scores to {OUTPUT}...")
    news.to_parquet(OUTPUT)

    rows_6h = evaluate(news["TOKS_total"].to_numpy(),
                       news["is_sepsis_6h"].to_numpy(),
                       "is_sepsis_6h")
    rows_12h = evaluate(news["TOKS_total"].to_numpy(),
                        news["is_sepsis_12h"].to_numpy(),
                        "is_sepsis_12h")

    print(f"\nAppending results to {EXCEL_PATH}...")
    append_to_excel(rows_6h, "6h Sepsis Prediction", "Danish TOKS 2.1 proxy (Nemati 2018 replication, 4h)")
    append_to_excel(rows_12h, "12h Sepsis Prediction", "Danish TOKS 2.1 proxy (Nemati 2018 replication, 4h)")
    print("Done.")


if __name__ == "__main__":
    main()
