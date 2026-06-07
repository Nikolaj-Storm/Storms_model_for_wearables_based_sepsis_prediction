# Copyright (c) 2026 Nikolaj Storm Petersen. Licensed under CC BY-NC 4.0.
# Non-commercial use only. If you use or adapt this code, please cite the author.
# See LICENSE and CITATION.cff  |  https://creativecommons.org/licenses/by-nc/4.0/

# ============================================================================
#  build_combined.py
#  Stage: 6 - Visualization / appendix
#
#  PURPOSE
#    Merge benchmark results from several model tiers (dummy, raw vitals,
#    engineered single models, V13b stacked ensemble) into a single styled
#    Excel workbook with one sheet per prediction horizon (6h, 12h) plus a
#    legend sheet documenting metrics, model tiers, and operating points.
#
#  INPUTS
#    /PATH/TO/INPUT/dummy.xlsx                              (dummy classifier results)
#    /PATH/TO/INPUT/raw vital only - base algorythms.xlsx   (raw-vitals LR/RF/XGB)
#    /PATH/TO/INPUT/optimised_performance_v2.csv            (engineered single models)
#    /PATH/TO/INPUT/v13b_final_report.xlsx                  (V13b stacked ensemble)
#  OUTPUTS
#    /PATH/TO/PROJECT/V13b_Combined_Results.xlsx
#
#  USER-EDITABLE SETTINGS  (grep the body for the tag  EDIT:  to find each)
#    dummy.xlsx path           -  input workbook with stratified dummy results
#    raw vital only path       -  input workbook with raw-vitals base algorithms
#    optimised_performance_v2  -  input CSV with engineered single-model results
#    v13b_final_report.xlsx    -  input workbook with V13b ensemble results
#    output .xlsx path         -  where the combined workbook is written
#    No hardcoded metric constants in this script; all numbers are read from
#    the input files above.
#
#  REQUIRES: openpyxl, csv (stdlib)
# ============================================================================

import csv, openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def pct(s):
    if s is None: return None
    if isinstance(s, (float, int)): return float(s) / 100 if float(s) > 1.5 else float(s)
    return float(str(s).replace('%','').strip()) / 100

def macro_f1(tp, fp, fn, tn):
    tp, fp, fn, tn = int(tp), int(fp), int(fn), int(tn)
    f1p = 2*tp / (2*tp + fp + fn) if (2*tp + fp + fn) else 0
    f1n = 2*tn / (2*tn + fn + fp) if (2*tn + fn + fp) else 0
    return round((f1p + f1n) / 2, 4)

def bs(): return Side(style='thin')
def bdr(): return Border(left=bs(), right=bs(), top=bs(), bottom=bs())

COLS = [
    ("Model Tier",                 18),
    ("Algorithm / Model",          28),
    ("Variant / Operating Point",  28),
    ("Threshold",                   9),
    ("AUPRC",                       9),
    ("Macro F1",                    9),
    ("Precision",                   9),
    ("Recall",                     10),
    ("F1 (pos class)",             13),
    ("ROC AUC",                    10),
    ("TP",                          8),
    ("FN",                          8),
    ("FP",                          8),
    ("TN",                          9),
    ("TP %",                        8),
    ("FN %",                        8),
    ("FP %",                        8),
    ("TN %",                        8),
    ("Train Acc",                  10),
    ("Test Acc",                   10),
]
COL_NAMES  = [c[0] for c in COLS]
N_COLS = len(COLS)
RESOLUTIONS = ['15-minute', '1-hour', '4-hour']

# ── parsers ───────────────────────────────────────────────────────────────────
def parse_dummy():
    # EDIT: input workbook with stratified dummy classifier results
    wb = load_workbook('/PATH/TO/INPUT/dummy.xlsx')
    ws = wb.active
    res_map = {'15_min':'15-minute','1_hour':'1-hour','4_hour':'4-hour'}
    recs = []
    for r in ws.iter_rows(values_only=True):
        if r[0] and 'Dummy' in str(r[0]):
            res_raw = str(r[0]).split('(')[0].strip()
            tp,fn,fp,tn = int(r[10]),int(r[11]),int(r[12]),int(r[13])
            recs.append({"Resolution": res_map.get(res_raw, res_raw),
                "Model Tier":"Dummy Classifier","Algorithm / Model":"Stratified Dummy",
                "Variant / Operating Point":"Default threshold (0.50)","Threshold":0.50,
                "AUPRC":r[9],"Macro F1":macro_f1(tp,fp,fn,tn),
                "Precision":r[3],"Recall":r[4],"F1 (pos class)":r[6],"ROC AUC":r[8],
                "TP":tp,"FN":fn,"FP":fp,"TN":tn,
                "TP %":pct(r[14]),"FN %":pct(r[15]),"FP %":pct(r[16]),"TN %":pct(r[17]),
                "Train Acc":pct(r[1]),"Test Acc":pct(r[2])})
    return recs

def parse_raw():
    # EDIT: input workbook with raw-vitals base algorithms (LR/RF/XGB)
    wb = load_workbook('/PATH/TO/INPUT/raw vital only - base algorythms.xlsx')
    ws = wb.active
    res_map = {'15_min':'15-minute','1_hour':'1-hour','4_hour':'4-hour'}
    alg_labels = {'XGBoost':'XGBoost','Random Forest':'Random Forest','Logistic Regression':'Logistic Regression'}
    recs = []
    for r in ws.iter_rows(values_only=True):
        if r[0] and any(x in str(r[0]) for x in alg_labels):
            parts = str(r[0]).split('(')
            res = res_map.get(parts[0].strip(), parts[0].strip())
            algo = parts[1].replace(')','').strip() if len(parts)>1 else parts[0]
            tp,fn,fp,tn = int(r[10]),int(r[11]),int(r[12]),int(r[13])
            recs.append({"Resolution":res,"Model Tier":"Raw Vitals (No Feat. Eng.)",
                "Algorithm / Model":algo,"Variant / Operating Point":"Default threshold (0.50)","Threshold":0.50,
                "AUPRC":r[9],"Macro F1":macro_f1(tp,fp,fn,tn),
                "Precision":r[3],"Recall":r[4],"F1 (pos class)":r[6],"ROC AUC":r[8],
                "TP":tp,"FN":fn,"FP":fp,"TN":tn,
                "TP %":pct(r[14]),"FN %":pct(r[15]),"FP %":pct(r[16]),"TN %":pct(r[17]),
                "Train Acc":pct(r[1]),"Test Acc":pct(r[2])})
    return recs

def parse_opt():
    res_map = {'15_min':'15-minute','1_hour':'1-hour','4_hour':'4-hour'}
    alg_map = {'RF Optimised':'Random Forest','LR Optimised':'Logistic Regression','XGB Optimised':'XGBoost'}
    recs = []
    # EDIT: input CSV with engineered single-model (optimised) results
    with open('/PATH/TO/INPUT/optimised_performance_v2.csv') as f:
        for row in csv.DictReader(f):
            raw = row['Resolution']
            parts = raw.split('(')
            res = res_map.get(parts[0].strip(), parts[0].strip())
            full_algo = parts[1].replace(')','').strip() if len(parts)>1 else parts[0]
            if 'thr=' in full_algo:
                algo_key = full_algo.split('thr=')[0].strip()
                variant = 'Custom threshold (0.300)'; thr = 0.300
            else:
                algo_key = full_algo; variant = 'Default threshold (0.50)'; thr = 0.50
            algo = alg_map.get(algo_key, algo_key)
            tp,fn,fp,tn = int(row['TP (Count)']),int(row['FN (Count)']),int(row['FP (Count)']),int(row['TN (Count)'])
            recs.append({"Resolution":res,"Model Tier":"Engineered Features (Single Model)",
                "Algorithm / Model":algo,"Variant / Operating Point":variant,"Threshold":thr,
                "AUPRC":float(row['AUPRC']),"Macro F1":macro_f1(tp,fp,fn,tn),
                "Precision":float(row['Precision']),"Recall":float(row['Recall']),
                "F1 (pos class)":float(row['F1 Score']),"ROC AUC":float(row['ROC AUC']),
                "TP":tp,"FN":fn,"FP":fp,"TN":tn,
                "TP %":pct(row['TP (%)']),"FN %":pct(row['FN (%)']),"FP %":pct(row['FP (%)']),"TN %":pct(row['TN (%)']),
                "Train Acc":pct(row['Train Accuracy (%)']),"Test Acc":pct(row['Test Accuracy (%)'])})
    return recs

def parse_v13b(sheet_name):
    # EDIT: input workbook with V13b stacked-ensemble final results
    wb = load_workbook('/PATH/TO/INPUT/v13b_final_report.xlsx')
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    res_clean = {'15-minute resolution':'15-minute','1-hour resolution':'1-hour','4-hour resolution':'4-hour'}
    op_labels  = {'F1-max (balanced)':'F1-maximising threshold (primary)',
                  'Max sens, FPR<0.50':'Max sensitivity, FPR < 50%',
                  'Max TP, FPR<0.20':'Max TP, FPR < 20%'}
    recs = []; cur_res = None; in_data = False
    for r in rows:
        if r[0] and 'resolution' in str(r[0]).lower():
            cur_res = res_clean.get(str(r[0]).lower().strip(), r[0]); in_data = False
        elif r[0] == 'Meta-learner':
            in_data = True
        elif in_data and r[0] in ('LR','XGB') and r[12] is not None:
            algo = 'Logistic Regression (meta)' if r[0]=='LR' else 'XGBoost (meta)'
            tp,fn,fp,tn = int(r[13]),int(r[14]),int(r[15]),int(r[16])
            recs.append({"Resolution":cur_res,"Model Tier":"V13b Stacked Ensemble",
                "Algorithm / Model":algo,"Variant / Operating Point":op_labels.get(r[1],r[1]),"Threshold":r[2],
                "AUPRC":r[12],"Macro F1":r[10],"Precision":r[5],"Recall":r[6],
                "F1 (pos class)":r[8],"ROC AUC":r[11],
                "TP":tp,"FN":fn,"FP":fp,"TN":tn,
                "TP %":r[17]/100,"FN %":r[18]/100,"FP %":r[19]/100,"TN %":r[20]/100,
                "Train Acc":r[3]/100,"Test Acc":r[4]/100})
    return recs

DUMMY = parse_dummy()
RAW   = parse_raw()
OPT   = parse_opt()
V6    = parse_v13b('6h prediction')
V12   = parse_v13b('12h prediction')

ALL_6H  = DUMMY + RAW + OPT + V6
ALL_12H = V12

# Verify resolution distribution
for res in RESOLUTIONS:
    n6  = len([r for r in ALL_6H  if r['Resolution'] == res])
    n12 = len([r for r in ALL_12H if r['Resolution'] == res])
    print(f"{res}: 6h={n6}, 12h={n12}")

# ── workbook builder ──────────────────────────────────────────────────────────
SEC_FILLS = {
    '15-minute': PatternFill("solid", fgColor="1F3864"),
    '1-hour':    PatternFill("solid", fgColor="1F3864"),
    '4-hour':    PatternFill("solid", fgColor="1F3864"),
}

def apply_col_widths(ws):
    for i, (name, w) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_title(ws, text, row):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=14, name="Arial", color="1F3864")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
    ws.row_dimensions[row].height = 24

def write_section_label(ws, text, row):
    label = f"  {text}"
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(bold=True, size=10, name="Arial", color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F3864")
    c.alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
    ws.row_dimensions[row].height = 20

def write_table(ws, records, table_name, start_row, style_name="TableStyleMedium2"):
    """Write header + data rows, then register as an Excel Table."""
    # header row
    for col_idx, col_name in enumerate(COL_NAMES, start=1):
        c = ws.cell(row=start_row, column=col_idx, value=col_name)
        c.font = Font(bold=True, size=9, name="Arial")
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[start_row].height = 28

    # data rows
    end_row = start_row
    for rec in records:
        end_row += 1
        ws.row_dimensions[end_row].height = 15
        for col_idx, col_name in enumerate(COL_NAMES, start=1):
            val = rec.get(col_name)
            c = ws.cell(row=end_row, column=col_idx, value=val)
            c.font = Font(size=9, name="Arial")
            c.alignment = Alignment(horizontal='center', vertical='center')
            # number formats
            if col_name in ("AUPRC","Macro F1","Precision","Recall","F1 (pos class)","ROC AUC","Threshold"):
                c.number_format = '0.0000'
            elif col_name in ("TP %","FN %","FP %","TN %","Train Acc","Test Acc"):
                c.number_format = '0.00%'
            elif col_name in ("TP","FN","FP","TN"):
                c.number_format = '#,##0'

    # Excel Table object
    ref = f"A{start_row}:{get_column_letter(N_COLS)}{end_row}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tbl)
    return end_row + 1  # return next free row

def build_sheet(ws, all_records, sheet_title, horizon_label):
    apply_col_widths(ws)
    write_title(ws, sheet_title, row=1)
    ws.row_dimensions[2].height = 8

    current_row = 3
    tbl_counter = 0
    for res in RESOLUTIONS:
        recs = [r for r in all_records if r['Resolution'] == res]
        if not recs:
            continue
        label_text = f"{res.upper()} RESOLUTION  ({len(recs)} model configurations)"
        write_section_label(ws, label_text, current_row)
        current_row += 1

        tbl_counter += 1
        safe_name = f"T_{horizon_label}_{res.replace('-','').replace(' ','_')}_{tbl_counter}"
        current_row = write_table(ws, recs, safe_name, current_row)
        current_row += 2  # blank gap

    ws.freeze_panes = 'A3'

# ── create workbook ───────────────────────────────────────────────────────────
wb_out = Workbook()
ws6  = wb_out.active
ws6.title  = "6h Sepsis Prediction"
ws12 = wb_out.create_sheet("12h Sepsis Prediction")

build_sheet(ws6,  ALL_6H,  "6-Hour Sepsis-3 Prediction Horizon  --  All Model Tiers", "6h")
build_sheet(ws12, ALL_12H, "12-Hour Sepsis-3 Prediction Horizon  --  V13b Stacked Ensemble", "12h")

# ── legend sheet ──────────────────────────────────────────────────────────────
wsl = wb_out.create_sheet("Legend")
wsl.column_dimensions['A'].width = 34
wsl.column_dimensions['B'].width = 78

legend = [
    ("METRIC NOTES", None),
    ("AUPRC", "Area Under Precision-Recall Curve. PRIMARY METRIC 1. Random baseline = class prevalence (~0.018 at 1-hr/6h). Higher is better."),
    ("Macro F1", "Macro-averaged F1 = (F1_pos + F1_neg) / 2. PRIMARY METRIC 2. Random baseline = 0.50. Values below 0.50 are operationally worse than random."),
    ("ROC AUC", "Included for completeness only. NOT a primary metric -- misleading under severe class imbalance (~2% positive class)."),
    ("Precision", "TP / (TP + FP). Fraction of sepsis alerts that are correct."),
    ("Recall", "TP / (TP + FN). Fraction of true sepsis cases flagged (sensitivity)."),
    ("F1 (pos class)", "Harmonic mean of Precision and Recall for the positive (sepsis) class only."),
    (None, None),
    ("MODEL TIER DESCRIPTIONS", None),
    ("Dummy Classifier", "Scikit-learn stratified dummy. Samples labels according to class prevalence. Establishes the floor any useful model must exceed."),
    ("Raw Vitals (No Feat. Eng.)", "LR / RF / XGB trained on 6 raw wearable vital signs only (HR, RR, SpO2, SBP, DBP, Temperature). No derived features."),
    ("Engineered Features (Single Model)", "LR / RF / XGB trained on the full 90-feature set including CUSUM drift stats, EMA trends, rolling std, shock index, and co-occurrence flags."),
    ("V13b Stacked Ensemble", "Final champion model. Etiology-specific NOSE submodels (sepsis, cardiovascular, respiratory, renal, metabolic) feeding an XGBoost or LR meta-learner."),
    (None, None),
    ("OPERATING POINTS (V13b)", None),
    ("F1-maximising threshold (primary)", "Threshold selected to maximise Macro F1. Primary clinical reference point used in the organizational analysis."),
    ("Max sensitivity, FPR < 50%",        "Highest achievable sensitivity while keeping false positive rate below 50%."),
    ("Max TP, FPR < 20%",                 "Maximum true positive count while keeping false positive rate below 20%."),
    (None, None),
    ("MACRO F1 NOTE", None),
    ("Computation for non-V13b tiers",
     "Macro F1 is computed directly from TP/FP/FN/TN counts using the formula: "
     "F1_pos = 2TP/(2TP+FP+FN), F1_neg = 2TN/(2TN+FN+FP), Macro F1 = (F1_pos+F1_neg)/2. "
     "V13b rows use the value stored in the results file."),
]

h_font = Font(bold=True, size=10, name="Arial", color="FFFFFF")
h_fill = PatternFill("solid", fgColor="1F3864")
b_font = Font(bold=True, size=9, name="Arial")
n_font = Font(size=9, name="Arial")

for i, (key, val) in enumerate(legend, start=1):
    ca = wsl.cell(row=i, column=1, value=key)
    cb = wsl.cell(row=i, column=2, value=val)
    ca.alignment = Alignment(wrap_text=True, vertical='top')
    cb.alignment = Alignment(wrap_text=True, vertical='top')
    wsl.row_dimensions[i].height = 30
    if key and key == key.upper() and key.replace(' ','').isalpha() or (key and key.endswith(("NOTES","TIONS","POINT","NOTE","TIERS"))):
        ca.font = h_font; ca.fill = h_fill
        wsl.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        wsl.row_dimensions[i].height = 20
    elif key:
        ca.font = b_font; cb.font = n_font

# EDIT: output path for the combined results workbook
out = '/PATH/TO/PROJECT/V13b_Combined_Results.xlsx'
wb_out.save(out)
print(f"Saved: {out}")
